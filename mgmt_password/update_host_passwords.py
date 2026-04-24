#!/usr/bin/env python3

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font

REQUIRED_HEADERS = (
    "IP",
    "관리명",
    "분야",
    "OS/애플리케이션",
    "계정형식",
    "사용자명",
    "password",
    "become",
    "become_password",
)
MATCH_KEYS = frozenset({"분야", "OS/애플리케이션", "계정형식", "사용자명"})
PATTERN_TYPES = frozenset(
    {
        "manager_name",
        "ip_all",
        "ip_all_padded",
        "ip_octet_1",
        "ip_octet_2",
        "ip_octet_3",
        "ip_octet_4",
        "ip_octet_1_padded",
        "ip_octet_2_padded",
        "ip_octet_3_padded",
        "ip_octet_4_padded",
    }
)
TRUTHY_STRINGS = frozenset({"1", "true", "t", "yes", "y", "on"})
PATTERN_SLICE_LIMITS = {
    "manager_name": 50,
    "ip_all": 12,
    "ip_all_padded": 12,
    "ip_octet_1": 3,
    "ip_octet_2": 3,
    "ip_octet_3": 3,
    "ip_octet_4": 3,
    "ip_octet_1_padded": 3,
    "ip_octet_2_padded": 3,
    "ip_octet_3_padded": 3,
    "ip_octet_4_padded": 3,
}


class ValidationError(ValueError):
    pass


@dataclass(frozen=True)
class PatternSpec:
    pattern_type: str
    prefix: str
    suffix: str
    start: Optional[int]
    end: Optional[int]


@dataclass(frozen=True)
class RuleSpec:
    target: str
    index: int
    match: Dict[str, Tuple[str, ...]]
    pattern: PatternSpec


@dataclass(frozen=True)
class UpdateSummary:
    total_rows: int
    password_updates: int
    become_password_updates: int
    updated_rows: int
    not_updated_rows: int
    password_rule_hits: Tuple[int, ...]
    become_password_rule_hits: Tuple[int, ...]
    output_path: Path


@dataclass(frozen=True)
class ResultRow:
    values: Dict[str, Any]
    updated: bool


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update password fields in a host Excel file.")
    parser.add_argument("--config", required=True, help="Path to the JSON config file")
    return parser.parse_args(argv)


def load_json_config(config_path: Path) -> Mapping[str, Any]:
    try:
        with config_path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except json.JSONDecodeError as exc:
        raise ValidationError(f"Invalid JSON in config file: {exc}") from exc

    if not isinstance(payload, dict):
        raise ValidationError("Config root must be a JSON object")
    return payload


def resolve_config_path(base_dir: Path, raw_path: Any, field_name: str) -> Path:
    if not isinstance(raw_path, str) or not raw_path.strip():
        raise ValidationError(f"{field_name} must be a non-empty string")

    candidate = Path(raw_path)
    if not candidate.is_absolute():
        candidate = base_dir / candidate
    return candidate


def normalize_match_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def parse_match_candidates(raw_value: Any, key: str, context: str) -> Tuple[str, ...]:
    if isinstance(raw_value, list):
        if not raw_value:
            raise ValidationError(f"{context} match.{key} array must not be empty")

        candidates: List[str] = []
        for item in raw_value:
            if isinstance(item, (list, dict)):
                raise ValidationError(f"{context} match.{key} array items must be scalar values")
            candidates.append(normalize_match_value(item))
        return tuple(candidates)

    if isinstance(raw_value, dict):
        raise ValidationError(f"{context} match.{key} must be a scalar value or an array")

    return (normalize_match_value(raw_value),)


def parse_positive_int(value: Any, field_name: str, context: str) -> int:
    if isinstance(value, bool):
        raise ValidationError(f"{context} {field_name} must be a positive integer")
    if isinstance(value, int):
        parsed = value
    elif isinstance(value, str) and value.isdigit():
        parsed = int(value)
    else:
        raise ValidationError(f"{context} {field_name} must be a positive integer")

    if parsed < 1:
        raise ValidationError(f"{context} {field_name} must be at least 1")
    return parsed


def parse_pattern(raw_pattern: Any, context: str) -> PatternSpec:
    if not isinstance(raw_pattern, dict):
        raise ValidationError(f"{context} pattern must be an object")

    pattern_type = raw_pattern.get("type")
    if pattern_type not in PATTERN_TYPES:
        allowed = ", ".join(sorted(PATTERN_TYPES))
        raise ValidationError(f"{context} pattern.type must be one of: {allowed}")

    prefix = raw_pattern.get("prefix")
    if not isinstance(prefix, str) or not prefix:
        raise ValidationError(f"{context} pattern.prefix must be a non-empty string")

    raw_suffix = raw_pattern.get("suffix", "")
    if raw_suffix is None:
        suffix = ""
    elif isinstance(raw_suffix, str):
        suffix = raw_suffix
    else:
        raise ValidationError(f"{context} pattern.suffix must be a string when provided")

    raw_start = raw_pattern.get("start")
    raw_end = raw_pattern.get("end")
    has_start = raw_start is not None
    has_end = raw_end is not None

    if has_start != has_end:
        raise ValidationError(f"{context} pattern.start and pattern.end must be provided together")

    start = None
    end = None
    if has_start and has_end:
        start = parse_positive_int(raw_start, "pattern.start", context)
        end = parse_positive_int(raw_end, "pattern.end", context)
        if start > end:
            raise ValidationError(f"{context} pattern.start must be less than or equal to pattern.end")
        slice_limit = PATTERN_SLICE_LIMITS[pattern_type]
        if start > slice_limit:
            raise ValidationError(
                f"{context} pattern.start must be between 1 and {slice_limit} for {pattern_type}"
            )
        if end > slice_limit:
            raise ValidationError(
                f"{context} pattern.end must be between 1 and {slice_limit} for {pattern_type}"
            )

    return PatternSpec(
        pattern_type=pattern_type,
        prefix=prefix,
        suffix=suffix,
        start=start,
        end=end,
    )


def parse_rules(raw_rules: Any, target: str) -> List[RuleSpec]:
    if raw_rules is None:
        return []
    if not isinstance(raw_rules, list):
        raise ValidationError(f"{target}_rules must be an array")

    parsed_rules: List[RuleSpec] = []
    for index, raw_rule in enumerate(raw_rules, start=1):
        context = f"{target}_rules[{index}]"
        if not isinstance(raw_rule, dict):
            raise ValidationError(f"{context} must be an object")

        raw_match = raw_rule.get("match", {})
        if raw_match is None:
            raw_match = {}
        if not isinstance(raw_match, dict):
            raise ValidationError(f"{context} match must be an object")

        unknown_keys = sorted(set(raw_match) - MATCH_KEYS)
        if unknown_keys:
            raise ValidationError(f"{context} match has unsupported keys: {', '.join(unknown_keys)}")

        match = {
            key: parse_match_candidates(value, key, context)
            for key, value in raw_match.items()
        }
        pattern = parse_pattern(raw_rule.get("pattern"), context)
        parsed_rules.append(RuleSpec(target=target, index=index, match=match, pattern=pattern))

    return parsed_rules


def ensure_output_path(input_path: Path, output_path: Path) -> None:
    resolved_input = input_path.resolve(strict=False)
    resolved_output = output_path.resolve(strict=False)
    if resolved_input == resolved_output:
        raise ValidationError("output_excel must be different from input_excel")
    if output_path.exists():
        raise ValidationError(f"Output file already exists: {output_path}")


def get_header_map(worksheet: Any) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column_index).value
        if value is None:
            continue
        header_map[str(value)] = column_index

    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_map]
    if missing_headers:
        raise ValidationError(f"Missing required headers: {', '.join(missing_headers)}")

    return header_map


def normalize_truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value == 1
    if isinstance(value, str):
        return value.strip().lower() in TRUTHY_STRINGS
    return False


def parse_ipv4_parts(raw_ip: Any) -> List[str]:
    if raw_ip is None:
        raise ValidationError("IP value is required for IP-based patterns")

    text = str(raw_ip).strip()
    parts = text.split(".")
    if len(parts) != 4:
        raise ValidationError(f"Invalid IPv4 address: {text}")

    normalized_parts: List[str] = []
    for part in parts:
        if not part.isdigit():
            raise ValidationError(f"Invalid IPv4 address: {text}")
        value = int(part)
        if value < 0 or value > 255:
            raise ValidationError(f"Invalid IPv4 address: {text}")
        normalized_parts.append(str(value))
    return normalized_parts


def apply_slice(value: str, pattern: PatternSpec, row_number: int) -> str:
    if pattern.start is None or pattern.end is None:
        return value
    if pattern.end > len(value):
        raise ValidationError(
            f"Row {row_number}: pattern slice {pattern.start}-{pattern.end} exceeds source length {len(value)}"
        )
    return value[pattern.start - 1 : pattern.end]


def build_source_value(row_values: Mapping[str, Any], pattern: PatternSpec, row_number: int) -> str:
    if pattern.pattern_type == "manager_name":
        return apply_slice(normalize_match_value(row_values["관리명"]), pattern, row_number)

    ip_parts = parse_ipv4_parts(row_values["IP"])
    if pattern.pattern_type == "ip_all":
        return apply_slice("".join(ip_parts), pattern, row_number)
    if pattern.pattern_type == "ip_all_padded":
        return apply_slice("".join(part.zfill(3) for part in ip_parts), pattern, row_number)

    octet_index_map = {
        "ip_octet_1": 0,
        "ip_octet_2": 1,
        "ip_octet_3": 2,
        "ip_octet_4": 3,
        "ip_octet_1_padded": 0,
        "ip_octet_2_padded": 1,
        "ip_octet_3_padded": 2,
        "ip_octet_4_padded": 3,
    }
    octet_index = octet_index_map[pattern.pattern_type]
    octet_value = ip_parts[octet_index]
    if pattern.pattern_type.endswith("_padded"):
        octet_value = octet_value.zfill(3)
    return apply_slice(octet_value, pattern, row_number)


def render_password(row_values: Mapping[str, Any], pattern: PatternSpec, row_number: int) -> str:
    source_value = build_source_value(row_values, pattern, row_number)
    return f"{pattern.prefix}{source_value}{pattern.suffix}"


def row_matches(row_values: Mapping[str, Any], match: Mapping[str, Tuple[str, ...]]) -> bool:
    for key, expected_values in match.items():
        if normalize_match_value(row_values.get(key)) not in expected_values:
            return False
    return True


def is_blank_data_row(row_values: Mapping[str, Any]) -> bool:
    return all(value is None for value in row_values.values())


def find_matching_rule(row_values: Mapping[str, Any], rules: Sequence[RuleSpec]) -> Optional[RuleSpec]:
    for rule in rules:
        if row_matches(row_values, rule.match):
            return rule
    return None


def read_row_values(worksheet: Any, row_number: int, header_map: Mapping[str, int]) -> Dict[str, Any]:
    return {
        header: worksheet.cell(row=row_number, column=column_index).value
        for header, column_index in header_map.items()
    }


def create_update_result_sheet(
    workbook: Any,
    ordered_headers: Sequence[str],
    result_rows: Sequence[ResultRow],
) -> None:
    result_sheet = workbook.create_sheet(title="업데이트 결과", index=1)
    result_headers = list(ordered_headers) + ["업데이트됨"]
    result_sheet.append(result_headers)
    for column_index in range(1, len(result_headers) + 1):
        result_sheet.cell(row=1, column=column_index).font = Font(bold=True)
    result_sheet.freeze_panes = "A2"

    for result_row in result_rows:
        row_values = [result_row.values.get(header) for header in ordered_headers]
        row_values.append("됨" if result_row.updated else "안됨")
        result_sheet.append(row_values)


def print_summary(summary: UpdateSummary) -> None:
    print(f"output: {summary.output_path}")
    print(f"rows: {summary.total_rows}")
    print(f"password_updates: {summary.password_updates}")
    print(f"become_password_updates: {summary.become_password_updates}")
    print(f"updated_rows: {summary.updated_rows}")
    print(f"not_updated_rows: {summary.not_updated_rows}")
    for index, hits in enumerate(summary.password_rule_hits, start=1):
        print(f"password_rules[{index}]: {hits}")
    for index, hits in enumerate(summary.become_password_rule_hits, start=1):
        print(f"become_password_rules[{index}]: {hits}")


def run_update_from_config(config_path: Path) -> UpdateSummary:
    resolved_config_path = config_path.resolve(strict=False)
    payload = load_json_config(resolved_config_path)
    base_dir = resolved_config_path.parent

    input_path = resolve_config_path(base_dir, payload.get("input_excel"), "input_excel")
    output_path = resolve_config_path(base_dir, payload.get("output_excel"), "output_excel")
    if not input_path.exists():
        raise ValidationError(f"Input file does not exist: {input_path}")

    ensure_output_path(input_path, output_path)
    password_rules = parse_rules(payload.get("password_rules"), "password")
    become_password_rules = parse_rules(payload.get("become_password_rules"), "become_password")

    workbook = load_workbook(input_path)
    try:
        if not workbook.sheetnames:
            raise ValidationError("Workbook does not contain any sheets")

        worksheet = workbook[workbook.sheetnames[0]]
        header_map = get_header_map(worksheet)
        password_rule_hits = [0 for _ in password_rules]
        become_password_rule_hits = [0 for _ in become_password_rules]
        password_updates = 0
        become_password_updates = 0
        total_rows = 0
        updated_rows = 0
        not_updated_rows = 0
        ordered_headers = list(header_map.keys())
        result_rows: List[ResultRow] = []

        for row_number in range(2, worksheet.max_row + 1):
            row_values = read_row_values(worksheet, row_number, header_map)
            if is_blank_data_row(row_values):
                continue

            total_rows += 1
            final_row_values = dict(row_values)
            row_updated = False

            password_rule = find_matching_rule(row_values, password_rules)
            if password_rule is not None:
                password_rule_hits[password_rule.index - 1] += 1
                new_password = render_password(row_values, password_rule.pattern, row_number)
                password_cell = worksheet.cell(row=row_number, column=header_map["password"])
                final_row_values["password"] = new_password
                if password_cell.value != new_password:
                    password_cell.value = new_password
                    password_updates += 1
                    row_updated = True

            if normalize_truthy(row_values["become"]):
                become_password_rule = find_matching_rule(row_values, become_password_rules)
                if become_password_rule is not None:
                    become_password_rule_hits[become_password_rule.index - 1] += 1
                    new_become_password = render_password(row_values, become_password_rule.pattern, row_number)
                    become_password_cell = worksheet.cell(row=row_number, column=header_map["become_password"])
                    final_row_values["become_password"] = new_become_password
                    if become_password_cell.value != new_become_password:
                        become_password_cell.value = new_become_password
                        become_password_updates += 1
                        row_updated = True

            if row_updated:
                updated_rows += 1
            else:
                not_updated_rows += 1
            result_rows.append(ResultRow(values=final_row_values, updated=row_updated))

        create_update_result_sheet(workbook, ordered_headers, result_rows)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()

    return UpdateSummary(
        total_rows=total_rows,
        password_updates=password_updates,
        become_password_updates=become_password_updates,
        updated_rows=updated_rows,
        not_updated_rows=not_updated_rows,
        password_rule_hits=tuple(password_rule_hits),
        become_password_rule_hits=tuple(become_password_rule_hits),
        output_path=output_path,
    )


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        summary = run_update_from_config(Path(args.config))
    except ValidationError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print_summary(summary)
    return 0


if __name__ == "__main__":
    sys.exit(main())
