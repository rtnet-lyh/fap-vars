import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from mgmt_password.update_host_passwords import (
    PatternSpec,
    ValidationError,
    normalize_truthy,
    parse_rules,
    render_password,
    run_update_from_config,
)


HEADERS = [
    "IP",
    "관리명",
    "설명",
    "인벤토리",
    "그룹",
    "장비중요도",
    "호스트 변수",
    "분야",
    "OS/애플리케이션",
    "제품명",
    "패밀리",
    "버전",
    "계정형식",
    "사용자명",
    "password",
    "en_password",
    "become",
    "become_method",
    "become_user",
    "become_password",
    "연결 타입",
    "네트워크 OS",
    "cert_validation",
    "use_ssl",
    "포트",
]


def make_row(
    ip,
    manager_name,
    area,
    os_name,
    account_type,
    username,
    password,
    become,
    become_password,
):
    return [
        ip,
        manager_name,
        "desc",
        "inventory",
        "group",
        "high",
        None,
        area,
        os_name,
        "product",
        None,
        None,
        account_type,
        username,
        password,
        None,
        become,
        "su -",
        "root",
        become_password,
        "ssh",
        None,
        None,
        None,
        22,
    ]


class PatternHelpersTest(unittest.TestCase):
    def test_render_password_supports_manager_name_and_ip_patterns(self) -> None:
        row_values = {
            "IP": "192.168.1.123",
            "관리명": "testhost1",
        }

        manager_password = render_password(
            row_values,
            PatternSpec(pattern_type="manager_name", prefix="pw_", suffix="!", start=2, end=5),
            row_number=2,
        )
        ip_all_password = render_password(
            row_values,
            PatternSpec(pattern_type="ip_all", prefix="pw_", suffix="", start=None, end=None),
            row_number=2,
        )
        padded_password = render_password(
            row_values,
            PatternSpec(pattern_type="ip_all_padded", prefix="pw_", suffix="#", start=1, end=12),
            row_number=2,
        )
        octet_password = render_password(
            row_values,
            PatternSpec(pattern_type="ip_octet_2", prefix="pw_", suffix="", start=2, end=3),
            row_number=2,
        )
        padded_octet_password = render_password(
            row_values,
            PatternSpec(pattern_type="ip_octet_3_padded", prefix="pw_", suffix="", start=1, end=3),
            row_number=2,
        )

        self.assertEqual(manager_password, "pw_esth!")
        self.assertEqual(ip_all_password, "pw_1921681123")
        self.assertEqual(padded_password, "pw_192168001123#")
        self.assertEqual(octet_password, "pw_68")
        self.assertEqual(padded_octet_password, "pw_001")

    def test_parse_rules_requires_non_empty_prefix(self) -> None:
        with self.assertRaisesRegex(ValidationError, "pattern.prefix"):
            parse_rules(
                [
                    {
                        "match": {"분야": "서버"},
                        "pattern": {"type": "ip_all", "suffix": "!"},
                    }
                ],
                "password",
            )

        with self.assertRaisesRegex(ValidationError, "pattern.prefix"):
            parse_rules(
                [
                    {
                        "match": {"분야": "서버"},
                        "pattern": {"type": "ip_all", "prefix": ""},
                    }
                ],
                "password",
            )

    def test_parse_rules_accepts_slice_for_all_pattern_types(self) -> None:
        rules = parse_rules(
            [
                {"match": {}, "pattern": {"type": "manager_name", "prefix": "pw_", "start": 1, "end": 5}},
                {"match": {}, "pattern": {"type": "ip_all", "prefix": "pw_", "start": 1, "end": 12}},
                {"match": {}, "pattern": {"type": "ip_all_padded", "prefix": "pw_", "start": 1, "end": 12}},
                {"match": {}, "pattern": {"type": "ip_octet_1", "prefix": "pw_", "start": 1, "end": 3}},
                {"match": {}, "pattern": {"type": "ip_octet_1_padded", "prefix": "pw_", "start": 1, "end": 3}},
            ],
            "password",
        )

        self.assertEqual(rules[0].pattern.start, 1)
        self.assertEqual(rules[0].pattern.end, 5)
        self.assertEqual(rules[1].pattern.end, 12)
        self.assertEqual(rules[3].pattern.end, 3)

    def test_parse_rules_enforces_type_specific_slice_ranges(self) -> None:
        with self.assertRaisesRegex(ValidationError, "between 1 and 50"):
            parse_rules(
                [
                    {
                        "match": {},
                        "pattern": {"type": "manager_name", "prefix": "pw_", "start": 1, "end": 51},
                    }
                ],
                "password",
            )

        with self.assertRaisesRegex(ValidationError, "between 1 and 12"):
            parse_rules(
                [
                    {
                        "match": {},
                        "pattern": {"type": "ip_all", "prefix": "pw_", "start": 1, "end": 13},
                    }
                ],
                "password",
            )

        with self.assertRaisesRegex(ValidationError, "between 1 and 3"):
            parse_rules(
                [
                    {
                        "match": {},
                        "pattern": {"type": "ip_octet_1", "prefix": "pw_", "start": 1, "end": 4},
                    }
                ],
                "password",
            )

    def test_parse_rules_accepts_match_arrays(self) -> None:
        rules = parse_rules(
            [
                {
                    "match": {
                        "분야": ["서버", "네트워크"],
                        "OS/애플리케이션": ["LINUX", "UNIX"],
                        "계정형식": "SSH",
                        "사용자명": ["alpha", "beta"],
                    },
                    "pattern": {"type": "ip_all", "prefix": "pw_"},
                }
            ],
            "password",
        )

        self.assertEqual(rules[0].match["분야"], ("서버", "네트워크"))
        self.assertEqual(rules[0].match["OS/애플리케이션"], ("LINUX", "UNIX"))
        self.assertEqual(rules[0].match["계정형식"], ("SSH",))
        self.assertEqual(rules[0].match["사용자명"], ("alpha", "beta"))

    def test_parse_rules_rejects_empty_match_arrays(self) -> None:
        with self.assertRaisesRegex(ValidationError, "must not be empty"):
            parse_rules(
                [
                    {
                        "match": {"분야": []},
                        "pattern": {"type": "ip_all", "prefix": "pw_"},
                    }
                ],
                "password",
            )

    def test_normalize_truthy_accepts_excel_friendly_values(self) -> None:
        self.assertTrue(normalize_truthy(True))
        self.assertTrue(normalize_truthy("TRUE"))
        self.assertTrue(normalize_truthy(1))
        self.assertFalse(normalize_truthy(False))
        self.assertFalse(normalize_truthy("no"))
        self.assertFalse(normalize_truthy(None))


class WorkbookUpdateTest(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "호스트 파일 다운로드"
        worksheet.freeze_panes = "A2"
        worksheet.append(HEADERS)
        worksheet["A1"].font = Font(bold=True)
        worksheet.append(make_row("10.0.0.1", "hostalpha", "서버", "LINUX", "SSH", "alpha", "old1", True, "root1"))
        worksheet.append(make_row("10.0.0.2", "hostbeta", "서버", "UNIX", "SSH", "beta", "old2", True, "root2"))
        worksheet.append(make_row("10.0.0.3", "hostgamma", "서버", "LINUX", "SSH", "gamma", "old3", False, "root3"))
        worksheet.append(make_row("10.0.0.4", "hostdelta", "네트워크", "LINUX", "SSH", "delta", "old4", True, "root4"))
        workbook.save(path)
        workbook.close()

    def write_config(self, path: Path, payload) -> None:
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def test_run_update_applies_rules_and_preserves_sheet_properties(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_path = Path(tmp_dir)
            input_path = temp_path / "input.xlsx"
            output_path = temp_path / "output.xlsx"
            config_path = temp_path / "rules.json"
            self.create_workbook(input_path)
            self.write_config(
                config_path,
                {
                    "input_excel": input_path.name,
                    "output_excel": output_path.name,
                    "password_rules": [
                        {
                            "match": {
                                "분야": ["서버", "클라우드"],
                                "OS/애플리케이션": ["LINUX", "UNIX"],
                                "계정형식": ["SSH", "TELNET"],
                                "사용자명": ["alpha", "beta", "gamma"],
                            },
                            "pattern": {
                                "type": "manager_name",
                                "prefix": "pw_",
                                "start": 1,
                                "end": 4,
                                "suffix": "!",
                            },
                        },
                        {
                            "match": {"분야": "서버"},
                            "pattern": {
                                "type": "ip_all",
                                "prefix": "unused_",
                            },
                        },
                    ],
                    "become_password_rules": [
                        {
                            "match": {"분야": "서버", "OS/애플리케이션": "LINUX", "계정형식": "SSH"},
                            "pattern": {
                                "type": "ip_octet_4_padded",
                                "prefix": "root_",
                                "start": 2,
                                "end": 3,
                            },
                        }
                    ],
                },
            )

            summary = run_update_from_config(config_path)

            self.assertEqual(summary.total_rows, 4)
            self.assertEqual(summary.password_updates, 3)
            self.assertEqual(summary.become_password_updates, 1)
            self.assertEqual(summary.updated_rows, 3)
            self.assertEqual(summary.not_updated_rows, 1)
            self.assertEqual(summary.password_rule_hits, (3, 0))
            self.assertEqual(summary.become_password_rule_hits, (1,))

            workbook = load_workbook(output_path)
            worksheet = workbook[workbook.sheetnames[0]]
            result_sheet = workbook[workbook.sheetnames[1]]
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertTrue(worksheet["A1"].font.bold)
            self.assertEqual(worksheet["O2"].value, "pw_host!")
            self.assertEqual(worksheet["O3"].value, "pw_host!")
            self.assertEqual(worksheet["O4"].value, "pw_host!")
            self.assertEqual(worksheet["T2"].value, "root_01")
            self.assertEqual(worksheet["T3"].value, "root2")
            self.assertEqual(worksheet["T4"].value, "root3")
            self.assertEqual(worksheet["O5"].value, "old4")
            self.assertEqual(result_sheet.title, "업데이트 결과")
            self.assertEqual(result_sheet.freeze_panes, "A2")
            self.assertEqual(result_sheet.cell(row=1, column=26).value, "업데이트됨")
            self.assertEqual(result_sheet.cell(row=2, column=15).value, "pw_host!")
            self.assertEqual(result_sheet.cell(row=2, column=20).value, "root_01")
            self.assertEqual(result_sheet.cell(row=2, column=26).value, "됨")
            self.assertEqual(result_sheet.cell(row=3, column=26).value, "됨")
            self.assertEqual(result_sheet.cell(row=4, column=26).value, "됨")
            self.assertEqual(result_sheet.cell(row=5, column=26).value, "안됨")
            workbook.close()

    def test_run_update_rejects_existing_output_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_path = Path(tmp_dir)
            input_path = temp_path / "input.xlsx"
            output_path = temp_path / "output.xlsx"
            config_path = temp_path / "rules.json"
            self.create_workbook(input_path)
            output_path.write_text("exists", encoding="utf-8")
            self.write_config(
                config_path,
                {
                    "input_excel": input_path.name,
                    "output_excel": output_path.name,
                    "password_rules": [],
                    "become_password_rules": [],
                },
            )

            with self.assertRaisesRegex(ValidationError, "already exists"):
                run_update_from_config(config_path)

    def test_run_update_fails_when_slice_exceeds_source_length(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_path = Path(tmp_dir)
            input_path = temp_path / "input.xlsx"
            output_path = temp_path / "output.xlsx"
            config_path = temp_path / "rules.json"
            self.create_workbook(input_path)
            self.write_config(
                config_path,
                {
                    "input_excel": input_path.name,
                    "output_excel": output_path.name,
                    "password_rules": [
                        {
                            "match": {"사용자명": "alpha"},
                            "pattern": {
                                "type": "manager_name",
                                "prefix": "pw_",
                                "start": 1,
                                "end": 20,
                            },
                        }
                    ],
                    "become_password_rules": [],
                },
            )

            with self.assertRaisesRegex(ValidationError, "exceeds source length"):
                run_update_from_config(config_path)


if __name__ == "__main__":
    unittest.main()
