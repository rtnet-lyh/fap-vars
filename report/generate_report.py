#!/usr/bin/env python3

import argparse
import copy
import configparser
import html
import json
import os
import re
import ssl
import sys
import textwrap
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from abc import ABC, abstractmethod
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Set, Tuple, Type

DEFAULT_FAP_CONFIG_FILE = Path("/fap/ansible/conf/fap.conf")
DEFAULT_OUTPUT_ROOT = Path(__file__).resolve().parent / "output"
DEFAULT_PREVENTIVE_HWPX_TEMPLATE = (
    Path(__file__).resolve().parent / "templates" / "PREVENTIVE_CHECK_TEMPLATE .hwpx"
)
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')
MAX_SHEETS_PER_WORKBOOK = 250
HWPX_SECTION_PATH = "Contents/section0.xml"
PREVENTIVE_HWPX_CHECK_ITEM_WRAP_CHARS = 18
PREVENTIVE_HWPX_TABLE_ROW_BASE_HEIGHT = 282
PREVENTIVE_HWPX_TABLE_ROW_LINE_HEIGHT = 1200
PREVENTIVE_HWPX_TABLE_HEIGHT_PADDING = 2000
PREVENTIVE_HWPX_TABLE_PARAGRAPH_HEIGHT_PADDING = 566
HWPX_NAMESPACES = {
    "ha": "http://www.hancom.co.kr/hwpml/2011/app",
    "hp": "http://www.hancom.co.kr/hwpml/2011/paragraph",
    "hp10": "http://www.hancom.co.kr/hwpml/2016/paragraph",
    "hs": "http://www.hancom.co.kr/hwpml/2011/section",
    "hc": "http://www.hancom.co.kr/hwpml/2011/core",
    "hh": "http://www.hancom.co.kr/hwpml/2011/head",
    "hhs": "http://www.hancom.co.kr/hwpml/2011/history",
    "hm": "http://www.hancom.co.kr/hwpml/2011/master-page",
    "hpf": "http://www.hancom.co.kr/schema/2011/hpf",
    "dc": "http://purl.org/dc/elements/1.1/",
    "opf": "http://www.idpf.org/2007/opf/",
    "ooxmlchart": "http://www.hancom.co.kr/hwpml/2016/ooxmlchart",
    "hwpunitchar": "http://www.hancom.co.kr/hwpml/2016/HwpUnitChar",
    "epub": "http://www.idpf.org/2007/ops",
    "config": "urn:oasis:names:tc:opendocument:xmlns:config:1.0",
}
HP_NAMESPACE = HWPX_NAMESPACES["hp"]
HP_TAG = "{%s}" % HP_NAMESPACE

for _hwpx_prefix, _hwpx_namespace in HWPX_NAMESPACES.items():
    ET.register_namespace(_hwpx_prefix, _hwpx_namespace)


class SummaryRow(object):
    def __init__(
        self,
        job_id,
        category_type_name,
        run_status,
        started_time,
        finished_time,
        host_id,
        host_name,
        host_ip,
        host_status,
        total_items,
        vuln_items,
        error_items,
        score,
        host_started,
        host_finished,
        duration_sec,
        error_message,
    ):
        self.job_id = job_id
        self.category_type_name = category_type_name
        self.run_status = run_status
        self.started_time = started_time
        self.finished_time = finished_time
        self.host_id = host_id
        self.host_name = host_name
        self.host_ip = host_ip
        self.host_status = host_status
        self.total_items = total_items
        self.vuln_items = vuln_items
        self.error_items = error_items
        self.score = score
        self.host_started = host_started
        self.host_finished = host_finished
        self.duration_sec = duration_sec
        self.error_message = error_message

    @property
    def host_key(self):
        return (self.host_id, self.host_name, self.host_ip)

    @property
    def good_items(self) -> int:
        return self.vuln_items

    @property
    def vulnerable_items(self) -> int:
        return self.error_items

    @property
    def not_run_items(self) -> int:
        computed = self.total_items - self.vuln_items - self.error_items
        return computed if computed > 0 else 0

    @classmethod
    def from_mapping(cls, row):
        return cls(
            job_id=_to_int(row.get("job_id")),
            category_type_name=str(row.get("category_type_name") or ""),
            run_status=str(row.get("run_status") or ""),
            started_time=row.get("started_time"),
            finished_time=row.get("finished_time"),
            host_id=_to_int(row.get("host_id")),
            host_name=str(row.get("host_name") or ""),
            host_ip=str(row.get("host_ip") or ""),
            host_status=str(row.get("host_status") or ""),
            total_items=_to_int(row.get("total_items")),
            vuln_items=_to_int(row.get("vuln_items")),
            error_items=_to_int(row.get("error_items")),
            score=_to_float(row.get("score")),
            host_started=row.get("host_started"),
            host_finished=row.get("host_finished"),
            duration_sec=_to_optional_int(row.get("duration_sec")),
            error_message=str(row.get("error_message") or ""),
        )


class DetailRow(object):
    def __init__(
        self,
        job_id,
        host_id,
        host_name,
        host_ip,
        inspection_code,
        inspection_item_name,
        type_name,
        category_name,
        area_name,
        importance,
        is_required,
        application_type_name,
        application_name,
        application_version,
        result_status,
        message,
        raw_output,
        description,
        inspection_command,
        is_service_affect,
        action_content,
        checked_time,
    ):
        self.job_id = job_id
        self.host_id = host_id
        self.host_name = host_name
        self.host_ip = host_ip
        self.inspection_code = inspection_code
        self.inspection_item_name = inspection_item_name
        self.type_name = type_name
        self.category_name = category_name
        self.area_name = area_name
        self.importance = importance
        self.is_required = is_required
        self.application_type_name = application_type_name
        self.application_name = application_name
        self.application_version = application_version
        self.result_status = result_status
        self.message = message
        self.raw_output = raw_output
        self.description = description
        self.inspection_command = inspection_command
        self.is_service_affect = is_service_affect
        self.action_content = action_content
        self.checked_time = checked_time

    @property
    def host_key(self):
        return (self.host_id, self.host_name, self.host_ip)

    @classmethod
    def from_mapping(cls, row):
        return cls(
            job_id=_to_int(row.get("job_id")),
            host_id=_to_int(row.get("host_id")),
            host_name=str(row.get("host_name") or ""),
            host_ip=str(row.get("host_ip") or ""),
            inspection_code=str(row.get("inspection_code") or ""),
            inspection_item_name=str(row.get("inspection_item_name") or ""),
            type_name=str(row.get("type_name") or ""),
            category_name=str(row.get("category_name") or ""),
            area_name=str(row.get("area_name") or ""),
            importance=str(row.get("importance") or ""),
            is_required=_to_optional_bool(row.get("is_required")),
            application_type_name=str(row.get("application_type_name") or ""),
            application_name=str(row.get("application_name") or ""),
            application_version=str(row.get("application_version") or ""),
            result_status=str(row.get("result_status") or ""),
            message=str(row.get("message") or ""),
            raw_output=str(row.get("raw_output") or ""),
            description=str(row.get("description") or ""),
            inspection_command=str(row.get("inspection_command") or ""),
            is_service_affect=str(row.get("is_service_affect") or ""),
            action_content=str(row.get("action_content") or ""),
            checked_time=row.get("checked_time"),
        )


class OverviewMetrics(object):
    def __init__(
        self,
        job_id,
        category_type_name,
        average_score,
        average_total_items,
        average_good_items,
        average_vuln_items,
        average_not_run_items,
        target_count,
        type_count,
    ):
        self.job_id = job_id
        self.category_type_name = category_type_name
        self.average_score = average_score
        self.average_total_items = average_total_items
        self.average_good_items = average_good_items
        self.average_vuln_items = average_vuln_items
        self.average_not_run_items = average_not_run_items
        self.target_count = target_count
        self.type_count = type_count


class BaseReportGenerator(ABC):
    aliases: Sequence[str] = ()

    @abstractmethod
    def build_workbook(
        self,
        summary_rows: Sequence[SummaryRow],
        detail_rows: Sequence[DetailRow],
        overview_rows: Optional[Sequence[SummaryRow]] = None,
        overview_detail_rows: Optional[Sequence[DetailRow]] = None,
    ) -> Any:
        raise NotImplementedError

    def save_workbook(
        self,
        summary_rows: Sequence[SummaryRow],
        detail_rows: Sequence[DetailRow],
        output_path: Path,
        overview_rows: Optional[Sequence[SummaryRow]] = None,
        overview_detail_rows: Optional[Sequence[DetailRow]] = None,
    ) -> Path:
        workbook = self.build_workbook(
            summary_rows,
            detail_rows,
            overview_rows=overview_rows,
            overview_detail_rows=overview_detail_rows,
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def save_workbooks(
        self,
        summary_rows: Sequence[SummaryRow],
        detail_rows: Sequence[DetailRow],
        output_path: Path,
        max_sheets_per_workbook: int = MAX_SHEETS_PER_WORKBOOK,
    ):
        max_detail_sheets = max_sheets_per_workbook - 1
        if max_detail_sheets < 1:
            raise ValueError("max_sheets_per_workbook must be at least 2")

        unique_hosts = unique_summary_rows(summary_rows)
        if len(unique_hosts) <= max_detail_sheets:
            return [
                self.save_workbook(
                    summary_rows,
                    detail_rows,
                    output_path,
                    overview_rows=summary_rows,
                    overview_detail_rows=detail_rows,
                )
            ]

        saved_paths = []
        host_chunks = chunk_sequence(unique_hosts, max_detail_sheets)
        total_parts = len(host_chunks)
        for part_index, host_chunk in enumerate(host_chunks, start=1):
            host_keys = {summary_row.host_key for summary_row in host_chunk}
            part_summary_rows = [summary_row for summary_row in summary_rows if summary_row.host_key in host_keys]
            part_detail_rows = [detail_row for detail_row in detail_rows if detail_row.host_key in host_keys]
            part_output_path = build_split_output_path(output_path, part_index, total_parts)
            saved_paths.append(
                self.save_workbook(
                    part_summary_rows,
                    part_detail_rows,
                    part_output_path,
                    overview_rows=summary_rows,
                    overview_detail_rows=detail_rows,
                )
            )
        return saved_paths


class DefaultInspectionReportGenerator(BaseReportGenerator):
    aliases = ("inspection",)

    def build_workbook(
        self,
        summary_rows: Sequence[SummaryRow],
        detail_rows: Sequence[DetailRow],
        overview_rows: Optional[Sequence[SummaryRow]] = None,
        overview_detail_rows: Optional[Sequence[DetailRow]] = None,
    ) -> Any:
        workbook, styles, sheet_names, grouped_details, get_column_letter = build_report_workbook_base(
            summary_rows,
            detail_rows,
            overview_rows=overview_rows,
            overview_detail_rows=overview_detail_rows,
        )

        for summary_row in unique_summary_rows(summary_rows):
            detail_sheet = workbook.create_sheet(sheet_names[summary_row.host_key])
            render_default_detail_sheet(
                detail_sheet,
                summary_row,
                grouped_details.get(summary_row.host_key, []),
                styles,
                get_column_letter,
            )

        return workbook


class PreventiveInspectionReportGenerator(BaseReportGenerator):
    aliases = ("preventive-detail",)

    def build_workbook(
        self,
        summary_rows: Sequence[SummaryRow],
        detail_rows: Sequence[DetailRow],
        overview_rows: Optional[Sequence[SummaryRow]] = None,
        overview_detail_rows: Optional[Sequence[DetailRow]] = None,
    ) -> Any:
        workbook, styles, sheet_names, grouped_details, get_column_letter = build_report_workbook_base(
            summary_rows,
            detail_rows,
            overview_rows=overview_rows,
            overview_detail_rows=overview_detail_rows,
        )

        for summary_row in unique_summary_rows(summary_rows):
            detail_sheet = workbook.create_sheet(sheet_names[summary_row.host_key])
            render_preventive_detail_sheet(
                detail_sheet,
                summary_row,
                grouped_details.get(summary_row.host_key, []),
                styles,
                get_column_letter,
            )

        return workbook


class GovernmentChecklistReportGenerator(BaseReportGenerator):
    aliases = ("default", "government-checklist")

    def build_workbook(
        self,
        summary_rows: Sequence[SummaryRow],
        detail_rows: Sequence[DetailRow],
        overview_rows: Optional[Sequence[SummaryRow]] = None,
        overview_detail_rows: Optional[Sequence[DetailRow]] = None,
    ) -> Any:
        workbook, styles, sheet_names, grouped_details, get_column_letter = build_report_workbook_base(
            summary_rows,
            detail_rows,
            overview_rows=overview_rows,
            overview_detail_rows=overview_detail_rows,
            summary_sheet_renderer=build_government_checklist_summary_sheet,
        )

        for summary_row in unique_summary_rows(summary_rows):
            detail_sheet = workbook.create_sheet(sheet_names[summary_row.host_key])
            render_preventive_detail_sheet(
                detail_sheet,
                summary_row,
                grouped_details.get(summary_row.host_key, []),
                styles,
                get_column_letter,
            )

        return workbook


REPORT_GENERATORS = {}
for generator_class in (
    DefaultInspectionReportGenerator,
    PreventiveInspectionReportGenerator,
    GovernmentChecklistReportGenerator,
):
    for alias in generator_class.aliases:
        REPORT_GENERATORS[alias] = generator_class


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate an inspection Excel report for a job_id.")
    parser.add_argument("--job-id", type=int, required=True, help="Target inspection job_id.")
    parser.add_argument("--user-id", help="Optional user identifier for traceability.")
    parser.add_argument("--report-type", default="default", help="Registered report type. Default: default")
    parser.add_argument(
        "--output-format",
        choices=("xlsx", "docx"),
        default="xlsx",
        help=(
            "Output format. xlsx creates Excel reports; docx creates one document per inspection area. "
            "report-type=preventive always creates HWPX output."
        ),
    )
    parser.add_argument("--output-dir", help="Optional output root. A dated directory is created below it.")
    parser.add_argument("--output-name", default="점검보고서", help="Output filename prefix. Default: 점검보고서")
    parser.add_argument("--file-name", help="Optional exact output filename. When set, timestamp is not appended.")
    parser.add_argument(
        "--output-path",
        help=(
            "Optional explicit output path. If used with --file-name, this is treated as a directory path. "
            "Without --file-name, this is treated as a full output file path."
        ),
    )
    parser.add_argument(
        "--mock-host-count",
        type=int,
        default=0,
        help="Generate mock data for the given number of hosts instead of querying DB.",
    )
    parser.add_argument(
        "--mock-items-per-host",
        type=int,
        default=3,
        help="Number of mock detail rows to create per host when mock mode is enabled.",
    )
    return parser.parse_args(argv)


def resolve_user_id(cli_user_id: Optional[str]) -> Optional[str]:
    if cli_user_id:
        candidate = cli_user_id.strip()
        if candidate:
            return candidate

    for env_key in ("FAP_REPORT_USER_ID", "FAP_USER_ID", "REPORT_USER_ID", "USER_ID"):
        candidate = os.environ.get(env_key, "").strip()
        if candidate:
            return candidate

    return None


def load_api_config(path: Optional[Path] = None) -> Tuple[str, str]:
    config_path = path or DEFAULT_FAP_CONFIG_FILE
    if not config_path.exists():
        raise FileNotFoundError("API config file not found: %s" % config_path)

    parser = configparser.ConfigParser(interpolation=None)
    with config_path.open("r", encoding="utf-8") as file_handle:
        parser.read_file(file_handle)

    if not parser.has_section("API_SERVER"):
        raise ValueError("missing [API_SERVER] section in config file: %s" % config_path)

    api_url = parser.get("API_SERVER", "API_URL", fallback="").strip()
    api_token = parser.get("API_SERVER", "API_TOKEN", fallback="").strip()
    if not api_url or not api_token:
        raise ValueError("missing API_URL or API_TOKEN in config file: %s" % config_path)
    return api_url.rstrip("/"), api_token


def api_get_json(api_url: str, api_token: str, endpoint: str, params: Optional[Mapping[str, Any]] = None) -> Any:
    query = urllib.parse.urlencode(dict(params or {}))
    request_url = "%s%s" % (api_url, endpoint)
    if query:
        request_url = "%s?%s" % (request_url, query)

    request = urllib.request.Request(
        request_url,
        headers={
            "Cache-Control": "no-cache",
            "x-auth-token": "Bearer " + api_token,
            "Authorization": "Bearer " + api_token,
        },
        method="GET",
    )
    ssl_context = ssl._create_unverified_context()
    try:
        with urllib.request.urlopen(request, context=ssl_context, timeout=30) as response:
            payload = response.read().decode("utf-8")
            return json.loads(payload or "[]")
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", "ignore")
        raise RuntimeError("API request failed: status=%s url=%s body=%s" % (exc.code, request_url, body))
    except urllib.error.URLError as exc:
        raise RuntimeError("API request failed: url=%s error=%s" % (request_url, exc))
    except ValueError as exc:
        raise RuntimeError("invalid JSON response from API: url=%s error=%s" % (request_url, exc))


def extract_row_list(payload: Any, label: str) -> List[Mapping[str, Any]]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, Mapping)]
    if isinstance(payload, Mapping):
        for key in ("data", "rows", "items", "result", "content"):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, Mapping)]
    raise RuntimeError("%s API response must be a JSON array or wrapper containing an array" % label)


def fetch_report_rows(api_url: str, api_token: str, job_id: int, user_id: Optional[str] = None):
    params = {"jobId": job_id}
    if user_id:
        params["user_id"] = user_id

    summary_payload = api_get_json(
        api_url,
        api_token,
        "/api/system/getReportData/summary",
        params=params,
    )
    detail_payload = api_get_json(
        api_url,
        api_token,
        "/api/system/getReportData/detail",
        params=params,
    )

    summary_rows = filter_empty_summary_rows(
        [SummaryRow.from_mapping(row) for row in extract_row_list(summary_payload, "summary")]
    )
    detail_rows = [DetailRow.from_mapping(row) for row in extract_row_list(detail_payload, "detail")]
    return summary_rows, detail_rows


def filter_empty_summary_rows(summary_rows: Sequence[SummaryRow]) -> List[SummaryRow]:
    return [row for row in summary_rows if row.total_items > 0]


def get_report_generator(report_type: str) -> BaseReportGenerator:
    generator_class = REPORT_GENERATORS.get(report_type)
    if not generator_class:
        supported = ", ".join(sorted(REPORT_GENERATORS))
        raise ValueError(f"unsupported report type: {report_type} (supported: {supported})")
    return generator_class()


def build_mock_report_rows(job_id: int, host_count: int, items_per_host: int = 3):
    if host_count < 1:
        raise ValueError("mock_host_count must be greater than 0")
    if items_per_host < 1:
        raise ValueError("mock_items_per_host must be greater than 0")

    summary_rows = []
    detail_rows = []
    importance_cycle = ["1", "2", "3"]
    result_cycle = ["양호", "취약", "미실행"]

    for host_index in range(1, host_count + 1):
        host_name = f"MOCK-HOST-{host_index:03d}"
        host_ip = f"10.10.{(host_index - 1) // 250}.{(host_index - 1) % 250 + 1}"
        summary_rows.append(
            SummaryRow(
                job_id=job_id,
                category_type_name="Mock Linux",
                run_status="completed",
                started_time=datetime(2026, 3, 17, 9, 0, 0),
                finished_time=datetime(2026, 3, 17, 9, 30, 0),
                host_id=host_index,
                host_name=host_name,
                host_ip=host_ip,
                host_status="success",
                total_items=items_per_host,
                vuln_items=max(items_per_host - 2, 0),
                error_items=1 if items_per_host > 1 else 0,
                score=max(100 - host_index % 17, 70),
                host_started=datetime(2026, 3, 17, 9, 0, 0),
                host_finished=datetime(2026, 3, 17, 9, 5, 0),
                duration_sec=300,
                error_message="",
            )
        )

        for item_index in range(1, items_per_host + 1):
            detail_rows.append(
                DetailRow(
                    job_id=job_id,
                    host_id=host_index,
                    host_name=host_name,
                    host_ip=host_ip,
                    inspection_code=f"MOCK-{host_index:03d}-{item_index:03d}",
                    inspection_item_name=f"Mock Inspection Item {item_index}",
                    type_name=f"유형-{(item_index - 1) % 4 + 1}",
                    category_name=f"구분-{(item_index - 1) % 3 + 1}",
                    area_name=f"영역-{(item_index - 1) % 5 + 1}",
                    importance=importance_cycle[(item_index - 1) % len(importance_cycle)],
                    is_required=True,
                    application_type_name="mock-app-type",
                    application_name="mock-app",
                    application_version="1.0",
                    result_status=result_cycle[(item_index - 1) % len(result_cycle)],
                    message=f"Mock message for host {host_index}, item {item_index}",
                    raw_output=f"Mock raw output for host {host_index}, item {item_index}",
                    description=f"Mock description for host {host_index}, item {item_index}",
                    inspection_command=f"/usr/bin/mock-check --host {host_name} --item {item_index}",
                    is_service_affect="무" if item_index % 2 else "유",
                    action_content=f"Mock action content for host {host_index}, item {item_index}",
                    checked_time=datetime(2026, 3, 17, 9, 0, 0),
                )
            )

    return summary_rows, detail_rows


def compute_overview_metrics(
    summary_rows: Sequence[SummaryRow],
    detail_rows: Optional[Sequence[DetailRow]] = None,
) -> OverviewMetrics:
    first_row = summary_rows[0]
    unique_hosts = unique_summary_rows(summary_rows)
    distinct_type_names = {
        detail_row.type_name.strip()
        for detail_row in (detail_rows or [])
        if detail_row.type_name and detail_row.type_name.strip()
    }
    return OverviewMetrics(
        job_id=first_row.job_id,
        category_type_name=first_row.category_type_name,
        average_score=average_or_zero(row.score or 0 for row in summary_rows),
        average_total_items=average_or_zero(row.total_items for row in summary_rows),
        average_good_items=average_or_zero(row.good_items for row in summary_rows),
        average_vuln_items=average_or_zero(row.vulnerable_items for row in summary_rows),
        average_not_run_items=average_or_zero(row.not_run_items for row in summary_rows),
        target_count=len(unique_hosts),
        type_count=len(distinct_type_names),
    )


def build_detail_title_text(summary_row: SummaryRow, detail_rows: Sequence[DetailRow]) -> str:
    inspected_count = len(detail_rows)
    pass_statuses = {"pass", "양호", "success", "ok"}
    importance_totals = {"상": 0, "중": 0, "하": 0}
    importance_passes = {"상": 0, "중": 0, "하": 0}
    pass_count = 0

    for detail_row in detail_rows:
        normalized_result = (detail_row.result_status or "").strip().lower()
        importance_label = format_importance(detail_row.importance)
        if importance_label in importance_totals:
            importance_totals[importance_label] += 1
        if normalized_result in pass_statuses:
            pass_count += 1
            if importance_label in importance_passes:
                importance_passes[importance_label] += 1

    def rate_text(passed: int, total: int) -> str:
        if total < 1:
            return "-"
        return f"{format_metric((float(passed) / float(total)) * 100.0)}%"

    host_label = summary_row.host_name or summary_row.host_ip or "호스트"
    return (
        f"{host_label} 상세 점검    "
        f"점검 항목 개수: {inspected_count}    "
        f"전체 성공률: {rate_text(pass_count, inspected_count)}    "
        f"상/중/하 성공률: {rate_text(importance_passes['상'], importance_totals['상'])}/"
        f"{rate_text(importance_passes['중'], importance_totals['중'])}/"
        f"{rate_text(importance_passes['하'], importance_totals['하'])}"
    )


def build_detail_overview_text(detail_rows: Sequence[DetailRow]) -> str:
    if not detail_rows:
        return "유형 목록: -    영역 목록: -    중요도(상/중/하): 0/0/0    PASS: 0    FAIL: 0"

    type_names = sorted({detail_row.type_name.strip() for detail_row in detail_rows if detail_row.type_name.strip()})
    area_names = sorted({detail_row.area_name.strip() for detail_row in detail_rows if detail_row.area_name.strip()})

    importance_counts = {"상": 0, "중": 0, "하": 0}
    pass_count = 0
    fail_count = 0

    for detail_row in detail_rows:
        importance_label = format_importance(detail_row.importance)
        if importance_label in importance_counts:
            importance_counts[importance_label] += 1

        normalized_result = (detail_row.result_status or "").strip().lower()
        if normalized_result in {"pass", "양호", "success", "ok"}:
            pass_count += 1
        elif normalized_result in {"fail", "취약", "error", "failed"}:
            fail_count += 1

    type_text = ", ".join(type_names) if type_names else "-"
    area_text = ", ".join(area_names) if area_names else "-"
    return (
        f"유형 목록: {type_text}    "
        f"영역 목록: {area_text}    "
        f"중요도(상/중/하): {importance_counts['상']}/{importance_counts['중']}/{importance_counts['하']}    "
        f"PASS: {pass_count}    FAIL: {fail_count}"
    )


def build_report_styles() -> Dict[str, Any]:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    return {
        "title_font": Font(size=16, bold=True, color="FFFFFF"),
        "subtitle_font": Font(size=10, bold=True, color="17365D"),
        "section_title_font": Font(size=12, bold=True, color="FFFFFF"),
        "section_label_font": Font(size=10, bold=True, color="FFFFFF"),
        "section_value_font": Font(size=11, bold=True, color="1F1F1F"),
        "body_font": Font(size=10, color="1F1F1F"),
        "muted_font": Font(size=10, bold=True, color="5B6573"),
        "header_font": Font(bold=True, color="FFFFFF"),
        "link_font": Font(color="0563C1", underline="single"),
        "title_fill": PatternFill("solid", fgColor="17365D"),
        "subtitle_fill": PatternFill("solid", fgColor="DCE6F1"),
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "section_fill": PatternFill("solid", fgColor="2F75B5"),
        "value_fill": PatternFill("solid", fgColor="F7FBFF"),
        "odd_row_fill": PatternFill("solid", fgColor="FFFFFF"),
        "even_row_fill": PatternFill("solid", fgColor="F6F9FC"),
        "success_fill": PatternFill("solid", fgColor="DDEBF7"),
        "danger_fill": PatternFill("solid", fgColor="FDE9E7"),
        "warning_fill": PatternFill("solid", fgColor="FFF2CC"),
        "neutral_fill": PatternFill("solid", fgColor="EDEDED"),
        "importance_high_fill": PatternFill("solid", fgColor="F4CCCC"),
        "importance_medium_fill": PatternFill("solid", fgColor="FFF2CC"),
        "importance_low_fill": PatternFill("solid", fgColor="D9EAD3"),
        "top_alignment": Alignment(horizontal="left", vertical="center"),
        "center_alignment": Alignment(horizontal="center", vertical="center"),
        "left_alignment": Alignment(horizontal="left", vertical="center"),
        "detail_row_alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "detail_top_alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
        "section_border": Border(
            left=Side(style="thin", color="B8CCE4"),
            right=Side(style="thin", color="B8CCE4"),
            top=Side(style="thin", color="B8CCE4"),
            bottom=Side(style="thin", color="B8CCE4"),
        ),
    }


def resolve_state_fill(status_text: str, styles: Mapping[str, Any]) -> Any:
    normalized = (status_text or "").strip().lower()
    if normalized in {"success", "completed", "done", "ok", "pass", "양호"}:
        return styles["success_fill"]
    if normalized in {"fail", "failed", "error", "취약", "danger"}:
        return styles["danger_fill"]
    if normalized in {"running", "progress", "pending", "warning", "미실행"}:
        return styles["warning_fill"]
    return styles["neutral_fill"]


def resolve_importance_fill(importance_text: str, styles: Mapping[str, Any]) -> Any:
    normalized = (importance_text or "").strip()
    if normalized == "상":
        return styles["importance_high_fill"]
    if normalized == "중":
        return styles["importance_medium_fill"]
    if normalized == "하":
        return styles["importance_low_fill"]
    return styles["neutral_fill"]


def build_report_workbook_base(
    summary_rows: Sequence[SummaryRow],
    detail_rows: Sequence[DetailRow],
    *,
    overview_rows: Optional[Sequence[SummaryRow]] = None,
    overview_detail_rows: Optional[Sequence[DetailRow]] = None,
    summary_sheet_renderer: Optional[Any] = None,
):
    if not summary_rows:
        raise RuntimeError("summary query returned no rows")

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "요약"

    styles = build_report_styles()
    overview = compute_overview_metrics(overview_rows or summary_rows, overview_detail_rows or detail_rows)
    sheet_names = build_sheet_name_map(summary_rows)

    grouped_details = defaultdict(list)
    for detail in detail_rows:
        grouped_details[detail.host_key].append(detail)

    renderer = summary_sheet_renderer or build_summary_sheet
    renderer(summary_sheet, summary_rows, overview, sheet_names, styles, get_column_letter, grouped_details)

    return workbook, styles, sheet_names, grouped_details, get_column_letter


def build_summary_sheet(
    summary_sheet: Any,
    summary_rows: Sequence[SummaryRow],
    overview: OverviewMetrics,
    sheet_names: Mapping[Tuple[int, str, str], str],
    styles: Mapping[str, Any],
    get_column_letter: Any,
    grouped_details: Optional[Mapping[Tuple[int, str, str], Sequence[DetailRow]]] = None,
) -> None:
    from copy import copy

    summary_sheet.sheet_view.showGridLines = False

    summary_sheet.merge_cells("A1:M1")
    summary_sheet["A1"] = "점검 보고서"
    summary_sheet["A1"].font = styles["title_font"]
    summary_sheet["A1"].fill = styles["title_fill"]
    summary_sheet["A1"].alignment = styles["center_alignment"]
    summary_sheet.row_dimensions[1].height = 26

    summary_sheet.merge_cells("A2:M2")
    summary_sheet["A2"] = "작업 정보"
    summary_sheet["A2"].font = styles["subtitle_font"]
    summary_sheet["A2"].fill = styles["subtitle_fill"]
    summary_sheet["A2"].alignment = styles["left_alignment"]
    summary_sheet.row_dimensions[2].height = 20

    write_label_value_pairs_row(
        summary_sheet,
        row=3,
        pairs=[
            (2, 3, 4, "작업번호", overview.job_id),
            (6, 7, 7, "점검 대상", overview.target_count),
            (8, 9, 9, "점검 유형", overview.type_count),
        ],
        label_font=styles["section_label_font"],
        value_font=styles["section_value_font"],
        label_fill=styles["section_fill"],
        value_fill=styles["value_fill"],
        border=styles["section_border"],
        label_alignment=styles["center_alignment"],
        value_alignment=styles["center_alignment"],
    )
    summary_sheet.row_dimensions[3].height = 24

    summary_sheet.merge_cells("A4:M4")
    summary_sheet["A4"] = "평균 지표"
    summary_sheet["A4"].font = styles["subtitle_font"]
    summary_sheet["A4"].fill = styles["subtitle_fill"]
    summary_sheet["A4"].alignment = styles["left_alignment"]
    summary_sheet.row_dimensions[4].height = 20

    write_label_value_pairs_row(
        summary_sheet,
        row=5,
        pairs=[
            (2, 3, 3, "점수", round(float(overview.average_score), 1)),
            (4, 5, 5, "항목", overview.average_total_items),
            (6, 7, 7, "양호", overview.average_good_items),
            (8, 9, 9, "취약", overview.average_vuln_items),
            (10, 11, 11, "미실행", overview.average_not_run_items),
        ],
        label_font=styles["section_label_font"],
        value_font=styles["section_value_font"],
        label_fill=styles["section_fill"],
        value_fill=styles["value_fill"],
        border=styles["section_border"],
        label_alignment=styles["center_alignment"],
        value_alignment=styles["center_alignment"],
    )
    for coordinate in ("C5", "E5", "G5", "I5", "K5"):
        summary_sheet[coordinate].number_format = "0.0"
    summary_sheet.row_dimensions[5].height = 24

    headers = [
        "번호",
        "관리명",
        "IP",
        "유형",
        "점수",
        "작업상태",
        "항목",
        "양호",
        "취약",
        "미실행",
        "시작시간",
        "종료시간",
        "소요시간",
    ]
    header_row = 7
    for index, header in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=header_row, column=index, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_alignment"]
        cell.border = styles["section_border"]

    for row_index, summary_row in enumerate(summary_rows, start=header_row + 1):
        number = row_index - header_row
        row_fill = styles["even_row_fill"] if number % 2 == 0 else styles["odd_row_fill"]

        link_cell = summary_sheet.cell(row=row_index, column=1, value=number)
        link_cell.hyperlink = f"#'{sheet_names[summary_row.host_key]}'!A1"
        link_cell.style = "Hyperlink"
        link_cell.font = styles["link_font"]
        link_cell.alignment = styles["center_alignment"]
        link_cell.border = styles["section_border"]
        link_cell.fill = row_fill

        values = [
            summary_row.host_name,
            summary_row.host_ip,
            summary_row.category_type_name,
            round(float(summary_row.score or 0), 2),
            summary_row.host_status,
            summary_row.total_items,
            summary_row.good_items,
            summary_row.vulnerable_items,
            summary_row.not_run_items,
            format_datetime(summary_row.host_started),
            format_datetime(summary_row.host_finished),
            format_duration(summary_row.duration_sec),
        ]
        for column_offset, value in enumerate(values, start=2):
            cell = summary_sheet.cell(row=row_index, column=column_offset, value=value)
            cell.font = styles["body_font"]
            cell.alignment = (
                styles["center_alignment"] if column_offset in {5, 6, 7, 8, 9, 10, 13} else styles["left_alignment"]
            )
            if column_offset == 5:
                cell.number_format = "0.00"
                style = copy(cell._style)
                style.numFmtId = 2
                cell._style = style
            cell.border = styles["section_border"]
            cell.fill = resolve_state_fill(str(value), styles) if column_offset == 6 else row_fill
        summary_sheet.row_dimensions[row_index].height = 21

    summary_sheet.freeze_panes = "A8"
    summary_sheet.auto_filter.ref = f"A7:M{max(len(summary_rows) + 7, 7)}"
    autosize_sheet(
        summary_sheet,
        get_column_letter,
        fixed_widths={
            "A": 6.88,
            "B": 18,
            "C": 16,
            "D": 14,
            "E": 8,
            "F": 12,
            "G": 8,
            "H": 8,
            "I": 8,
            "J": 12.13,
            "K": 20,
            "L": 20,
            "M": 12,
        },
    )


def build_government_checklist_summary_sheet(
    summary_sheet: Any,
    summary_rows: Sequence[SummaryRow],
    overview: OverviewMetrics,
    sheet_names: Mapping[Tuple[int, str, str], str],
    styles: Mapping[str, Any],
    get_column_letter: Any,
    grouped_details: Optional[Mapping[Tuple[int, str, str], Sequence[DetailRow]]] = None,
) -> None:
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.page_setup.orientation = "portrait"
    summary_sheet.page_setup.paperSize = summary_sheet.PAPERSIZE_A4
    summary_sheet.page_setup.fitToWidth = 1
    summary_sheet.page_setup.fitToHeight = 0
    summary_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    summary_sheet.page_margins.left = 0.3
    summary_sheet.page_margins.right = 0.3
    summary_sheet.page_margins.top = 0.5
    summary_sheet.page_margins.bottom = 0.5

    detail_map = grouped_details or {}
    area_sections = build_government_checklist_area_sections(summary_rows, detail_map)
    current_row = 1
    for section_index, section in enumerate(area_sections, start=1):
        if section_index > 1:
            current_row += 2
        current_row = render_government_checklist_summary_section(
            summary_sheet,
            current_row,
            section["area_name"],
            section["host_entries"],
            sheet_names,
            styles,
        )

    fixed_widths = {
        "A": 7,
        "B": 34,
    }
    max_host_count = max((len(section["host_entries"]) for section in area_sections), default=1)
    for host_index in range(max_host_count):
        start_column = 3 + (host_index * 3)
        fixed_widths[get_column_letter(start_column)] = 8
        fixed_widths[get_column_letter(start_column + 1)] = 8
        fixed_widths[get_column_letter(start_column + 2)] = 18
    autosize_sheet(
        summary_sheet,
        get_column_letter,
        fixed_widths=fixed_widths,
    )


def render_government_checklist_summary_section(
    worksheet: Any,
    start_row: int,
    inspection_area_name: str,
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
    sheet_names: Mapping[Tuple[int, str, str], str],
    styles: Mapping[str, Any],
) -> int:
    area_name = (inspection_area_name or "").strip() or "점검 영역"
    host_labels = [build_host_label(summary_row) for summary_row, _details in host_entries]
    checked_time = first_present_value(*build_government_checklist_checked_time_candidates(host_entries))
    end_column = max(5, 2 + (len(host_entries) * 3))
    host_detail_indexes = [
        index_details_by_government_checklist_item(details_for_host) for _summary_row, details_for_host in host_entries
    ]
    item_entries = build_government_checklist_item_entries(host_entries)
    all_details_for_area = [detail for _summary_row, details_for_host in host_entries for detail in details_for_host]

    write_merged_row(
        worksheet,
        row=start_row,
        start_column=1,
        end_column=end_column,
        value=f"{area_name} 점검 일지",
        font=styles["title_font"],
        fill=styles["title_fill"],
        alignment=styles["center_alignment"],
        border=styles["section_border"],
    )
    if host_entries:
        worksheet.cell(row=start_row, column=1).hyperlink = f"#'{sheet_names[host_entries[0][0].host_key]}'!A1"
    worksheet.row_dimensions[start_row].height = 28
    current_row = start_row + 1

    for meta_text in (
        f"□ {area_name} 점검",
        f"○ 장비명: {', '.join(host_labels) if host_labels else '-'}",
        f"○ 점검 일자: {format_datetime(checked_time)}",
    ):
        write_merged_row(
            worksheet,
            row=current_row,
            start_column=1,
            end_column=end_column,
            value=meta_text,
            font=styles["body_font"],
            fill=styles["value_fill"],
            alignment=styles["left_alignment"],
            border=styles["section_border"],
        )
        worksheet.row_dimensions[current_row].height = 22
        current_row += 1

    table_header_row = current_row
    worksheet.merge_cells(start_row=table_header_row, start_column=1, end_row=table_header_row + 1, end_column=1)
    worksheet.merge_cells(start_row=table_header_row, start_column=2, end_row=table_header_row + 1, end_column=2)
    for column_index, header in ((1, "번호"), (2, "점검항목")):
        cell = worksheet.cell(row=table_header_row, column=column_index, value=header)
        apply_cell_style(cell, styles["header_font"], styles["header_fill"], styles["center_alignment"], styles["section_border"])

    for host_index, (summary_row, _details_for_host) in enumerate(host_entries):
        host_column = 3 + (host_index * 3)
        worksheet.merge_cells(
            start_row=table_header_row,
            start_column=host_column,
            end_row=table_header_row,
            end_column=host_column + 2,
        )
        host_cell = worksheet.cell(row=table_header_row, column=host_column, value=build_host_label(summary_row))
        apply_cell_style(
            host_cell,
            styles["header_font"],
            styles["header_fill"],
            styles["center_alignment"],
            styles["section_border"],
        )
        for offset, header in enumerate(("정상", "비정상", "비고")):
            cell = worksheet.cell(row=table_header_row + 1, column=host_column + offset, value=header)
            apply_cell_style(
                cell,
                styles["header_font"],
                styles["header_fill"],
                styles["center_alignment"],
                styles["section_border"],
            )

    for header_row in (table_header_row, table_header_row + 1):
        for column_index in range(1, end_column + 1):
            cell = worksheet.cell(row=header_row, column=column_index)
            cell.fill = styles["header_fill"]
            cell.border = styles["section_border"]
            cell.alignment = styles["center_alignment"]
    worksheet.row_dimensions[current_row].height = 24
    worksheet.row_dimensions[current_row + 1].height = 22
    current_row += 2

    if not item_entries:
        write_merged_row(
            worksheet,
            row=current_row,
            start_column=1,
            end_column=end_column,
            value="상세 데이터가 없습니다.",
            font=styles["body_font"],
            fill=styles["value_fill"],
            alignment=styles["left_alignment"],
            border=styles["section_border"],
        )
        worksheet.row_dimensions[current_row].height = 24
        current_row += 1
    else:
        for item_number, (item_key, item_label) in enumerate(item_entries, start=1):
            row_fill = styles["even_row_fill"] if item_number % 2 == 0 else styles["odd_row_fill"]
            row_values = [item_number, item_label]
            for host_detail_index in host_detail_indexes:
                detail_row = host_detail_index.get(item_key)
                row_values.extend(build_government_checklist_result_cells(detail_row))

            for column_index in range(1, end_column + 1):
                value = row_values[column_index - 1] if column_index <= len(row_values) else ""
                alignment = (
                    styles["center_alignment"]
                    if column_index == 1 or (column_index >= 3 and (column_index - 3) % 3 in {0, 1})
                    else styles["detail_row_alignment"]
                )
                cell = worksheet.cell(row=current_row, column=column_index, value=value)
                apply_cell_style(cell, styles["body_font"], row_fill, alignment, styles["section_border"])

            note_values = [
                str(row_values[column_index]).strip()
                for column_index in range(4, len(row_values), 3)
                if str(row_values[column_index]).strip()
            ]
            worksheet.row_dimensions[current_row].height = max(
                estimate_text_row_height(
                    item_label,
                    chars_per_line=24,
                    min_height=26,
                    max_height=96,
                ),
                *(
                    estimate_text_row_height(
                        note_value,
                        chars_per_line=14,
                        min_height=26,
                        max_height=96,
                    )
                    for note_value in note_values
                )
            )
            current_row += 1

    result_summary_text = build_government_checklist_result_summary(all_details_for_area)
    action_summary_text = build_government_checklist_action_summary(host_entries)
    for title_text, body_text in (
        ("□ 점검 결과 요약", result_summary_text),
        ("□ 주요 조치내역", action_summary_text),
    ):
        write_merged_row(
            worksheet,
            row=current_row,
            start_column=1,
            end_column=end_column,
            value=title_text,
            font=styles["subtitle_font"],
            fill=styles["subtitle_fill"],
            alignment=styles["left_alignment"],
            border=styles["section_border"],
        )
        worksheet.row_dimensions[current_row].height = 22
        current_row += 1

        write_merged_row(
            worksheet,
            row=current_row,
            start_column=1,
            end_column=end_column,
            value=body_text,
            font=styles["body_font"],
            fill=styles["value_fill"],
            alignment=styles["detail_top_alignment"],
            border=styles["section_border"],
        )
        worksheet.row_dimensions[current_row].height = estimate_text_row_height(
            body_text,
            chars_per_line=82,
            min_height=26,
            max_height=120,
        )
        current_row += 1

    signature_pairs = (
        (1, 2, min(3, end_column), "점검자", ""),
        (max(4, end_column - 1), end_column, end_column, "확인자", ""),
    )
    write_label_value_pairs_row(
        worksheet,
        row=current_row,
        pairs=signature_pairs,
        label_font=styles["section_label_font"],
        value_font=styles["body_font"],
        label_fill=styles["section_fill"],
        value_fill=styles["value_fill"],
        border=styles["section_border"],
        label_alignment=styles["center_alignment"],
        value_alignment=styles["center_alignment"],
    )
    worksheet.row_dimensions[current_row].height = 32

    return current_row + 1


def build_government_checklist_area_sections(
    summary_rows: Sequence[SummaryRow],
    detail_map: Mapping[Tuple[int, str, str], Sequence[DetailRow]],
) -> List[Dict[str, Any]]:
    unique_hosts = unique_summary_rows(summary_rows)
    sections = []
    sections_by_area = {}

    for summary_row in unique_hosts:
        grouped_details = group_details_by_area(detail_map.get(summary_row.host_key, []))
        for area_name, details_for_area in grouped_details:
            if area_name not in sections_by_area:
                section = {"area_name": area_name, "host_entries": []}
                sections_by_area[area_name] = section
                sections.append(section)
            sections_by_area[area_name]["host_entries"].append((summary_row, details_for_area))

    if not sections:
        return [{"area_name": "점검 영역", "host_entries": [(summary_row, []) for summary_row in unique_hosts]}]
    return sections


def group_details_by_area(details_for_host: Sequence[DetailRow]) -> List[Tuple[str, Sequence[DetailRow]]]:
    if not details_for_host:
        return []

    grouped = {}
    for detail_row in details_for_host:
        area_name = (detail_row.area_name or "").strip() or "점검 영역"
        grouped.setdefault(area_name, []).append(detail_row)
    return [(area_name, details) for area_name, details in grouped.items()]


def build_government_checklist_checked_time_candidates(
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
) -> List[Any]:
    values = []
    for summary_row, details_for_host in host_entries:
        values.extend([summary_row.host_started, summary_row.started_time])
        values.extend(detail.checked_time for detail in details_for_host if detail.checked_time not in (None, ""))
    return values


def build_government_checklist_item_entries(
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
) -> List[Tuple[str, str]]:
    entries = []
    seen_keys = set()
    for _summary_row, details_for_host in host_entries:
        for detail_row in details_for_host:
            item_key = build_government_checklist_item_key(detail_row)
            if item_key in seen_keys:
                continue
            seen_keys.add(item_key)
            entries.append((item_key, build_government_checklist_item_label(detail_row)))
    return entries


def index_details_by_government_checklist_item(details_for_host: Sequence[DetailRow]) -> Dict[str, DetailRow]:
    indexed = {}
    for detail_row in details_for_host:
        item_key = build_government_checklist_item_key(detail_row)
        if item_key not in indexed:
            indexed[item_key] = detail_row
    return indexed


def build_government_checklist_item_key(detail_row: DetailRow) -> str:
    return (detail_row.inspection_item_name or detail_row.inspection_code or "").strip()


def build_government_checklist_item_label(detail_row: DetailRow) -> str:
    return (detail_row.inspection_item_name or detail_row.inspection_code or "-").strip()


def build_government_checklist_result_cells(detail_row: Optional[DetailRow]) -> List[str]:
    if detail_row is None:
        return ["[ ]", "[ ]", "-"]

    status_type = classify_detail_result_status(detail_row.result_status)
    return [
        "[✔]" if status_type == "pass" else "[ ]",
        "[✔]" if status_type == "fail" else "[ ]",
        build_government_checklist_note(detail_row, status_type),
    ]


def build_host_label(summary_row: SummaryRow) -> str:
    return (summary_row.host_name or summary_row.host_ip or "-").strip()


def apply_cell_style(cell: Any, font: Any, fill: Any, alignment: Any, border: Any) -> None:
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border


def write_merged_row(
    worksheet: Any,
    *,
    row: int,
    start_column: int,
    end_column: int,
    value: Any,
    font: Any,
    fill: Any,
    alignment: Any,
    border: Any,
) -> None:
    for column_index in range(start_column, end_column + 1):
        cell = worksheet.cell(row=row, column=column_index)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border

    if start_column < end_column:
        worksheet.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=end_column)

    value_cell = worksheet.cell(row=row, column=start_column, value=value)
    value_cell.font = font
    value_cell.fill = fill
    value_cell.alignment = alignment
    value_cell.border = border


def classify_detail_result_status(result_status: Any) -> str:
    normalized = ("" if result_status is None else str(result_status)).strip().lower()
    if normalized in {"pass", "양호", "success", "ok"}:
        return "pass"
    if normalized in {"fail", "취약", "error", "failed", "warn", "warning"}:
        return "fail"
    return "unknown"


def build_government_checklist_note(detail_row: DetailRow, status_type: str) -> str:
    message_text = (detail_row.message or "").strip()
    result_text = (detail_row.result_status or "").strip()
    if status_type == "unknown" and result_text:
        return f"{message_text}\n결과: {result_text}" if message_text else f"결과: {result_text}"
    return message_text or "-"


def build_government_checklist_result_summary(details_for_host: Sequence[DetailRow]) -> str:
    pass_count = 0
    fail_count = 0
    unknown_count = 0
    for detail_row in details_for_host:
        status_type = classify_detail_result_status(detail_row.result_status)
        if status_type == "pass":
            pass_count += 1
        elif status_type == "fail":
            fail_count += 1
        else:
            unknown_count += 1
    return (
        f"총 {len(details_for_host)}건 / "
        f"정상 {pass_count}건 / "
        f"비정상 {fail_count}건 / "
        f"확인필요 {unknown_count}건"
    )


def build_government_checklist_action_summary(
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
) -> str:
    action_items = []
    for summary_row, details_for_host in host_entries:
        host_label = build_host_label(summary_row)
        for detail_row in details_for_host:
            action_text = str(detail_row.action_content or "").strip()
            if action_text:
                action_items.append(f"{host_label}: {action_text}")
    if not action_items:
        return "-"
    return "\n".join(f"{index}. {action_item}" for index, action_item in enumerate(action_items, start=1))


def first_present_value(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def write_detail_sheet_intro(
    detail_sheet: Any,
    summary_row: SummaryRow,
    details_for_host: Sequence[DetailRow],
    styles: Mapping[str, Any],
    merge_end_column: str = "H",
) -> None:
    detail_sheet.sheet_view.showGridLines = False
    detail_sheet["A1"] = "요약으로 돌아가기"
    detail_sheet["A1"].hyperlink = "#'요약'!A1"
    detail_sheet["A1"].style = "Hyperlink"
    detail_sheet["A1"].font = styles["link_font"]

    detail_sheet.merge_cells(f"A2:{merge_end_column}2")
    detail_sheet["A2"] = build_detail_title_text(summary_row, details_for_host)
    detail_sheet["A2"].font = styles["section_title_font"]
    detail_sheet["A2"].fill = styles["title_fill"]
    detail_sheet["A2"].alignment = styles["top_alignment"]
    detail_sheet.row_dimensions[2].height = 24

    detail_sheet.merge_cells(f"A3:{merge_end_column}3")
    detail_sheet["A3"] = build_detail_overview_text(details_for_host)
    detail_sheet["A3"].font = styles["muted_font"]
    detail_sheet["A3"].fill = styles["value_fill"]
    detail_sheet["A3"].alignment = styles["left_alignment"]
    detail_sheet["A3"].border = styles["section_border"]
    detail_sheet.row_dimensions[3].height = 36


def render_default_detail_sheet(
    detail_sheet: Any,
    summary_row: SummaryRow,
    details_for_host: Sequence[DetailRow],
    styles: Mapping[str, Any],
    get_column_letter: Any,
) -> None:
    write_detail_sheet_intro(detail_sheet, summary_row, details_for_host, styles, merge_end_column="I")

    detail_headers = ["유형", "영역", "구분", "중요도", "점검 코드", "항목", "결과", "메세지", "상세"]
    detail_header_row = 5
    for index, header in enumerate(detail_headers, start=1):
        cell = detail_sheet.cell(row=detail_header_row, column=index, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_alignment"]
        cell.border = styles["section_border"]

    if not details_for_host:
        detail_sheet.merge_cells("A6:I6")
        detail_sheet["A6"] = "상세 데이터가 없습니다."
        detail_sheet["A6"].alignment = styles["left_alignment"]
        detail_sheet["A6"].fill = styles["value_fill"]
        detail_sheet["A6"].border = styles["section_border"]
    else:
        for row_index, detail_row in enumerate(details_for_host, start=detail_header_row + 1):
            row_fill = styles["even_row_fill"] if (row_index - detail_header_row) % 2 == 0 else styles["odd_row_fill"]
            row_values = [
                detail_row.type_name,
                detail_row.area_name,
                detail_row.category_name,
                format_importance(detail_row.importance),
                detail_row.inspection_code,
                detail_row.inspection_item_name,
                detail_row.result_status,
                detail_row.message,
                detail_row.raw_output,
            ]
            for column_index, value in enumerate(row_values, start=1):
                cell = detail_sheet.cell(row=row_index, column=column_index, value=value)
                cell.font = styles["body_font"]
                cell.alignment = styles["detail_row_alignment"]
                cell.border = styles["section_border"]
                if column_index == 4:
                    cell.fill = resolve_importance_fill(str(value), styles)
                elif column_index == 7:
                    cell.fill = resolve_state_fill(str(value), styles)
                else:
                    cell.fill = row_fill
            detail_sheet.row_dimensions[row_index].height = 36

    detail_sheet.freeze_panes = "A6"
    detail_sheet.auto_filter.ref = f"A5:I{max(len(details_for_host) + 5, 5)}"
    autosize_sheet(
        detail_sheet,
        get_column_letter,
        fixed_widths={
            "A": 12,
            "B": 16,
            "C": 16,
            "D": 10,
            "E": 18,
            "F": 32,
            "G": 12,
            "H": 38,
            "I": 56,
        },
    )


def render_preventive_detail_sheet(
    detail_sheet: Any,
    summary_row: SummaryRow,
    details_for_host: Sequence[DetailRow],
    styles: Mapping[str, Any],
    get_column_letter: Any,
) -> None:
    write_detail_sheet_intro(detail_sheet, summary_row, details_for_host, styles)

    detail_start_row = 5
    if not details_for_host:
        detail_sheet.merge_cells("A5:H5")
        detail_sheet["A5"] = "상세 데이터가 없습니다."
        detail_sheet["A5"].font = styles["body_font"]
        detail_sheet["A5"].alignment = styles["left_alignment"]
        detail_sheet["A5"].fill = styles["value_fill"]
        detail_sheet["A5"].border = styles["section_border"]
        detail_sheet.row_dimensions[5].height = 24
    else:
        for detail_index, detail_row in enumerate(details_for_host):
            start_row = detail_start_row + (detail_index * 12)
            row_fill = styles["even_row_fill"] if detail_index % 2 else styles["odd_row_fill"]
            importance_text = format_importance(detail_row.importance)

            write_label_value_pairs_row(
                detail_sheet,
                row=start_row,
                pairs=[
                    (1, 2, 3, "유형", detail_row.type_name),
                    (4, 5, 6, "영역", detail_row.area_name),
                    (7, 8, 8, "구분", detail_row.category_name),
                ],
                label_font=styles["section_label_font"],
                value_font=styles["body_font"],
                label_fill=styles["section_fill"],
                value_fill=row_fill,
                border=styles["section_border"],
                label_alignment=styles["center_alignment"],
                value_alignment=styles["left_alignment"],
            )
            detail_sheet.row_dimensions[start_row].height = 24

            write_label_value_pairs_row(
                detail_sheet,
                row=start_row + 1,
                pairs=[
                    (1, 2, 3, "애플리케이션유형", detail_row.application_type_name),
                    (4, 5, 6, "애플리케이션명", detail_row.application_name),
                    (7, 8, 8, "버전", detail_row.application_version),
                ],
                label_font=styles["section_label_font"],
                value_font=styles["body_font"],
                label_fill=styles["section_fill"],
                value_fill=row_fill,
                border=styles["section_border"],
                label_alignment=styles["center_alignment"],
                value_alignment=styles["left_alignment"],
            )
            detail_sheet.row_dimensions[start_row + 1].height = 24

            write_label_value_pairs_row(
                detail_sheet,
                row=start_row + 2,
                pairs=[
                    (1, 2, 3, "점검결과", detail_row.result_status),
                    (4, 5, 6, "중요도", importance_text),
                    (7, 8, 8, "점검코드", detail_row.inspection_code),
                ],
                label_font=styles["section_label_font"],
                value_font=styles["body_font"],
                label_fill=styles["section_fill"],
                value_fill=row_fill,
                border=styles["section_border"],
                label_alignment=styles["center_alignment"],
                value_alignment=styles["center_alignment"],
            )
            detail_sheet.cell(row=start_row + 2, column=2).fill = resolve_state_fill(detail_row.result_status, styles)
            detail_sheet.cell(row=start_row + 2, column=5).fill = resolve_importance_fill(importance_text, styles)
            detail_sheet.row_dimensions[start_row + 2].height = 24

            for row_offset, label, value, dynamic_height in (
                (3, "점검항목", detail_row.inspection_item_name, False),
                (4, "명령어", detail_row.inspection_command, False),
                (5, "상세", detail_row.raw_output, True),
                (6, "메세지", detail_row.message, True),
                (7, "설명", detail_row.description, True),
                (8, "서비스 영향 유/무", detail_row.is_service_affect, False),
                (9, "조치내역", detail_row.action_content, True),
            ):
                row_number = start_row + row_offset
                write_label_value_pairs_row(
                    detail_sheet,
                    row=row_number,
                    pairs=[(1, 2, 8, label, value)],
                    label_font=styles["section_label_font"],
                    value_font=styles["body_font"],
                    label_fill=styles["section_fill"],
                    value_fill=row_fill,
                    border=styles["section_border"],
                    label_alignment=styles["center_alignment"],
                    value_alignment=styles["detail_top_alignment"],
                )
                detail_sheet.row_dimensions[row_number].height = (
                    estimate_text_row_height(value) if dynamic_height else 24
                )

    detail_sheet.freeze_panes = "A5"
    autosize_sheet(
        detail_sheet,
        get_column_letter,
        fixed_widths={
            "A": 18,
            "B": 16,
            "C": 16,
            "D": 18,
            "E": 16,
            "F": 16,
            "G": 12,
            "H": 18,
        },
    )


def build_sheet_name_map(summary_rows: Sequence[SummaryRow]):
    used_names = set()
    mapping = {}
    for index, summary_row in enumerate(unique_summary_rows(summary_rows), start=1):
        base_name = summary_row.host_name or summary_row.host_ip or f"host_{index}"
        mapping[summary_row.host_key] = normalize_sheet_name(base_name, used_names)
    return mapping


def unique_summary_rows(summary_rows: Sequence[SummaryRow]):
    unique_rows = []
    seen_host_keys = set()
    for summary_row in summary_rows:
        if summary_row.host_key in seen_host_keys:
            continue
        seen_host_keys.add(summary_row.host_key)
        unique_rows.append(summary_row)
    return unique_rows


def chunk_sequence(items: Sequence[Any], chunk_size: int):
    return [list(items[index : index + chunk_size]) for index in range(0, len(items), chunk_size)]


def normalize_sheet_name(raw_name: str, used_names: Optional[Set[str]] = None) -> str:
    used_names = used_names if used_names is not None else set()
    cleaned = INVALID_SHEET_CHARS.sub("_", raw_name or "").strip().strip("'")
    cleaned = re.sub(r"\s+", " ", cleaned)
    base_name = cleaned or "host"
    candidate = base_name[:31]
    suffix = 1

    while any(existing.lower() == candidate.lower() for existing in used_names):
        suffix_text = f"_{suffix}"
        candidate = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_names.add(candidate)
    return candidate


def build_output_path(output_name: str, output_dir: Optional[str] = None, suffix: str = ".xlsx") -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    date_dir = datetime.now().strftime("%Y%m%d")
    root = Path(output_dir).expanduser() if output_dir else DEFAULT_OUTPUT_ROOT
    target_dir = root / date_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    output_stem = normalize_output_name(output_name)
    return (target_dir / f"{output_stem}_{timestamp}{normalize_suffix(suffix)}").resolve()


def build_named_output_path(
    file_name: str,
    output_path: Optional[str] = None,
    output_dir: Optional[str] = None,
    suffix: str = ".xlsx",
) -> Path:
    normalized_suffix = normalize_suffix(suffix)
    candidate_name = Path(file_name).name.strip()
    if not candidate_name:
        candidate_name = f"점검보고서{normalized_suffix}"
    if Path(candidate_name).suffix.lower() != normalized_suffix:
        candidate_name = f"{Path(candidate_name).stem}{normalized_suffix}"

    if output_path:
        target_dir = Path(output_path).expanduser()
    else:
        date_dir = datetime.now().strftime("%Y%m%d")
        root = Path(output_dir).expanduser() if output_dir else DEFAULT_OUTPUT_ROOT
        target_dir = root / date_dir

    target_dir.mkdir(parents=True, exist_ok=True)
    return (target_dir / candidate_name).resolve()


def build_explicit_output_path(output_path: str, suffix: str = ".xlsx") -> Path:
    normalized_suffix = normalize_suffix(suffix)
    candidate = Path(output_path).expanduser()
    if candidate.suffix.lower() != normalized_suffix:
        candidate = candidate.with_suffix(normalized_suffix)
    candidate.parent.mkdir(parents=True, exist_ok=True)
    return candidate.resolve()


def build_split_output_path(output_path: Path, part_index: int, total_parts: int) -> Path:
    suffix_width = max(2, len(str(total_parts)))
    return output_path.with_name(
        f"{output_path.stem}_part{part_index:0{suffix_width}d}{output_path.suffix}"
    )


def build_zip_output_path(output_path: Path) -> Path:
    return output_path.with_suffix(".zip")


def normalize_suffix(suffix: str) -> str:
    candidate = (suffix or ".xlsx").strip().lower()
    return candidate if candidate.startswith(".") else f".{candidate}"


def build_zip_archive(saved_paths: Sequence[Path], archive_path: Path) -> Path:
    archive_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(str(archive_path), "w", compression=zipfile.ZIP_DEFLATED) as zip_handle:
        for saved_path in saved_paths:
            zip_handle.write(str(saved_path), arcname=saved_path.name)
    return archive_path.resolve()


def save_government_checklist_docx_reports(
    summary_rows: Sequence[SummaryRow],
    detail_rows: Sequence[DetailRow],
    output_path: Path,
) -> Path:
    if not summary_rows:
        raise RuntimeError("summary query returned no rows")

    grouped_details = defaultdict(list)
    for detail in detail_rows:
        grouped_details[detail.host_key].append(detail)

    area_sections = build_government_checklist_area_sections(summary_rows, grouped_details)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    document_dir = output_path.with_suffix("")
    document_dir.mkdir(parents=True, exist_ok=True)

    saved_paths = []
    used_names = set()
    for section_index, section in enumerate(area_sections, start=1):
        area_name = section["area_name"]
        filename = build_area_document_filename(output_path.stem, area_name, section_index, used_names)
        document_path = document_dir / filename
        document_bytes = build_government_checklist_docx(area_name, section["host_entries"])
        document_path.write_bytes(document_bytes)
        saved_paths.append(document_path)

    if len(saved_paths) == 1:
        return saved_paths[0].resolve()
    return build_zip_archive(saved_paths, build_zip_output_path(output_path))


def save_preventive_hwpx_reports(
    summary_rows: Sequence[SummaryRow],
    detail_rows: Sequence[DetailRow],
    output_path: Path,
    template_path: Path = DEFAULT_PREVENTIVE_HWPX_TEMPLATE,
) -> Path:
    if not summary_rows:
        raise RuntimeError("summary query returned no rows")

    grouped_details = defaultdict(list)
    for detail in detail_rows:
        grouped_details[detail.host_key].append(detail)

    area_sections = build_government_checklist_area_sections(summary_rows, grouped_details)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if len(area_sections) == 1:
        section = area_sections[0]
        output_path.write_bytes(
            build_preventive_hwpx(
                section["area_name"],
                section["host_entries"],
                template_path=template_path,
            )
        )
        return output_path.resolve()

    document_dir = output_path.with_suffix("")
    document_dir.mkdir(parents=True, exist_ok=True)

    saved_paths = []
    used_names = set()
    for section_index, section in enumerate(area_sections, start=1):
        area_name = section["area_name"]
        filename = build_area_document_filename(
            output_path.stem,
            area_name,
            section_index,
            used_names,
            suffix=".hwpx",
        )
        document_path = document_dir / filename
        document_path.write_bytes(
            build_preventive_hwpx(
                area_name,
                section["host_entries"],
                template_path=template_path,
            )
        )
        saved_paths.append(document_path)

    return build_zip_archive(saved_paths, build_zip_output_path(output_path))


def build_preventive_hwpx(
    area_name: str,
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
    template_path: Path = DEFAULT_PREVENTIVE_HWPX_TEMPLATE,
) -> bytes:
    if not template_path.exists():
        raise FileNotFoundError("preventive HWPX template not found: %s" % template_path)

    from io import BytesIO

    buffer = BytesIO()
    with zipfile.ZipFile(str(template_path), "r") as template_zip:
        section_xml = template_zip.read(HWPX_SECTION_PATH)
        rendered_section_xml = build_preventive_hwpx_section_xml(section_xml, area_name, host_entries)

        with zipfile.ZipFile(buffer, "w") as output_zip:
            for info in template_zip.infolist():
                data = (
                    rendered_section_xml
                    if info.filename == HWPX_SECTION_PATH
                    else template_zip.read(info.filename)
                )
                output_zip.writestr(info, data)

    return buffer.getvalue()


def build_preventive_hwpx_section_xml(
    template_section_xml: bytes,
    area_name: str,
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
) -> bytes:
    root = ET.fromstring(template_section_xml)
    template_children = list(root)
    for child in template_children:
        root.remove(child)

    entries = list(host_entries)
    if not entries:
        return serialize_hwpx_section(root)

    for host_index, (summary_row, details_for_host) in enumerate(entries):
        section_children = [copy.deepcopy(child) for child in template_children]
        if host_index > 0:
            mark_hwpx_section_page_break(section_children)

        render_preventive_hwpx_host_section(
            section_children,
            area_name,
            summary_row,
            details_for_host,
        )
        root.extend(section_children)

    return serialize_hwpx_section(root)


def serialize_hwpx_section(root: ET.Element) -> bytes:
    body = ET.tostring(root, encoding="utf-8", short_empty_elements=True)
    return b'<?xml version="1.0" encoding="UTF-8" standalone="yes" ?>' + body


def mark_hwpx_section_page_break(section_children: Sequence[ET.Element]) -> None:
    for child in section_children:
        if child.tag == HP_TAG + "p":
            child.set("pageBreak", "1")
            return


def render_preventive_hwpx_host_section(
    section_children: List[ET.Element],
    area_name: str,
    summary_row: SummaryRow,
    details_for_host: Sequence[DetailRow],
) -> None:
    checked_time = format_datetime(
        first_present_value(
            *build_government_checklist_checked_time_candidates([(summary_row, details_for_host)])
        )
    )
    pass_count, fail_count = count_detail_result_statuses(details_for_host)
    replacements = {
        "{{INSPECTION_AREA}}": (area_name or "").strip() or "점검 영역",
        "{{HOST_NAME}}": build_host_label(summary_row),
        "{{CHECK_DATE}}": checked_time,
        "{{IS_PASS_CNT}}": str(pass_count),
        "{{IS_FAILED_CNT}}": str(fail_count),
    }

    table_paragraph_index = find_preventive_hwpx_detail_table_paragraph_index(section_children)
    extra_row_count = render_preventive_hwpx_detail_rows(section_children, details_for_host)
    resize_preventive_hwpx_table_paragraph(section_children, table_paragraph_index)
    remove_hwpx_spacer_paragraphs_after_index(section_children, table_paragraph_index, extra_row_count)
    replace_hwpx_placeholders(section_children, replacements)


def render_preventive_hwpx_detail_rows(
    section_children: Sequence[ET.Element],
    details_for_host: Sequence[DetailRow],
) -> int:
    table = find_preventive_hwpx_detail_table(section_children)
    if table is None:
        return 0

    rows = table.findall(HP_TAG + "tr")
    data_row = None
    data_row_index = None
    for row_index, row in enumerate(list(table)):
        if row.tag == HP_TAG + "tr" and hwpx_element_contains_text(row, "{{NUM}}"):
            data_row = row
            data_row_index = row_index
            break
    if data_row is None or data_row_index is None:
        return 0

    table.remove(data_row)
    render_details = list(details_for_host) or [None]
    spacer_remove_count = max(len(render_details) - 1, 0)
    for detail_index, detail_row in enumerate(render_details, start=1):
        row = copy.deepcopy(data_row)
        set_hwpx_table_row_address(row, detail_index)
        if detail_row is None:
            replace_hwpx_placeholders(
                [row],
                {
                    "{{NUM}}": "",
                    "{{IS_PASS}}": " ",
                    "{{IS_FAILED}}": " ",
                },
            )
            set_hwpx_table_cell_wrapped_text(row, 1, "상세 데이터가 없습니다.")
            set_preventive_hwpx_table_row_height(row, 1)
            set_hwpx_table_cell_text(row, 4, "")
        else:
            status_type = classify_detail_result_status(detail_row.result_status)
            check_item_lines = wrap_preventive_hwpx_check_item(
                build_government_checklist_item_label(detail_row)
            )
            replace_hwpx_placeholders(
                [row],
                {
                    "{{NUM}}": str(detail_index),
                    "{{IS_PASS}}": "✔" if status_type == "pass" else " ",
                    "{{IS_FAILED}}": "✔" if status_type == "fail" else " ",
                },
            )
            set_hwpx_table_cell_wrapped_text(row, 1, check_item_lines)
            set_preventive_hwpx_table_row_height(row, len(check_item_lines))
            spacer_remove_count += max(len(check_item_lines) - 1, 0)
            set_hwpx_table_cell_text(row, 4, "")
        table.insert(data_row_index + detail_index - 1, row)

    table.set("rowCnt", str(max(len(rows) - 1, 0) + len(render_details)))
    resize_preventive_hwpx_table_height(table)
    return spacer_remove_count


def wrap_preventive_hwpx_check_item(value: Any) -> List[str]:
    text = " ".join(str(value if value is not None else "").split())
    if not text:
        return [""]

    wrapped_lines = []
    for raw_line in text.splitlines() or [""]:
        wrapped_lines.extend(
            textwrap.wrap(
                raw_line,
                width=PREVENTIVE_HWPX_CHECK_ITEM_WRAP_CHARS,
                break_long_words=True,
                break_on_hyphens=False,
            )
            or [""]
        )
    return wrapped_lines or [""]


def set_hwpx_table_cell_wrapped_text(row: ET.Element, column_index: int, value: Any) -> None:
    lines = list(value) if isinstance(value, (list, tuple)) else wrap_preventive_hwpx_check_item(value)
    cells = row.findall(HP_TAG + "tc")
    if column_index >= len(cells):
        return

    cell = cells[column_index]
    sub_list = cell.find(HP_TAG + "subList")
    if sub_list is None:
        set_hwpx_table_cell_text(row, column_index, "\n".join(lines))
        return

    sub_list_children = list(sub_list)
    paragraphs = [child for child in sub_list_children if child.tag == HP_TAG + "p"]
    if not paragraphs:
        set_hwpx_table_cell_text(row, column_index, "\n".join(lines))
        return

    first_paragraph_index = sub_list_children.index(paragraphs[0])
    paragraph_template = paragraphs[0]
    for paragraph in paragraphs:
        sub_list.remove(paragraph)

    insert_index = first_paragraph_index
    for line in lines:
        paragraph = copy.deepcopy(paragraph_template)
        set_hwpx_paragraph_text(paragraph, line)
        sub_list.insert(insert_index, paragraph)
        insert_index += 1


def set_hwpx_paragraph_text(paragraph: ET.Element, value: Any) -> None:
    text_elements = list(paragraph.iter(HP_TAG + "t"))
    if text_elements:
        set_hwpx_text_element_plain_text(text_elements[0], value)
        for extra_text_element in text_elements[1:]:
            set_hwpx_text_element_plain_text(extra_text_element, "")
        return

    run = paragraph.find(".//" + HP_TAG + "run")
    if run is not None:
        text_element = ET.SubElement(run, HP_TAG + "t")
        text_element.text = "" if value is None else str(value)


def set_hwpx_text_element_plain_text(text_element: ET.Element, value: Any) -> None:
    for child in list(text_element):
        text_element.remove(child)
    text_element.text = "" if value is None else str(value)


def set_preventive_hwpx_table_row_height(row: ET.Element, line_count: int) -> None:
    row_height = estimate_preventive_hwpx_table_row_height(line_count)
    for cell in row.findall(HP_TAG + "tc"):
        cell_size = cell.find(HP_TAG + "cellSz")
        if cell_size is not None:
            cell_size.set("height", str(row_height))


def estimate_preventive_hwpx_table_row_height(line_count: int) -> int:
    extra_line_count = max(line_count - 1, 0)
    return PREVENTIVE_HWPX_TABLE_ROW_BASE_HEIGHT + (
        extra_line_count * PREVENTIVE_HWPX_TABLE_ROW_LINE_HEIGHT
    )


def resize_preventive_hwpx_table_height(table: ET.Element) -> None:
    table_size = table.find(HP_TAG + "sz")
    if table_size is None:
        return

    table_height = PREVENTIVE_HWPX_TABLE_HEIGHT_PADDING
    for row in table.findall(HP_TAG + "tr"):
        cell_heights = []
        for cell in row.findall(HP_TAG + "tc"):
            cell_size = cell.find(HP_TAG + "cellSz")
            if cell_size is not None:
                cell_heights.append(parse_hwpx_int(cell_size.get("height"), 0))
        table_height += max(cell_heights) if cell_heights else 0

    current_height = parse_hwpx_int(table_size.get("height"), 0)
    table_size.set("height", str(max(current_height, table_height)))


def resize_preventive_hwpx_table_paragraph(
    section_children: Sequence[ET.Element],
    paragraph_index: Optional[int],
) -> None:
    if paragraph_index is None or paragraph_index >= len(section_children):
        return

    paragraph = section_children[paragraph_index]
    table = find_preventive_hwpx_detail_table([paragraph])
    if table is None:
        return

    table_size = table.find(HP_TAG + "sz")
    if table_size is None:
        return

    paragraph_height = parse_hwpx_int(table_size.get("height"), 0) + PREVENTIVE_HWPX_TABLE_PARAGRAPH_HEIGHT_PADDING
    if paragraph_height < 1:
        return

    line_segment = paragraph.find(HP_TAG + "linesegarray/" + HP_TAG + "lineseg")
    if line_segment is None:
        return

    line_segment.set("vertsize", str(paragraph_height))
    line_segment.set("textheight", str(paragraph_height))
    line_segment.set("baseline", str(int(paragraph_height * 0.85)))


def parse_hwpx_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def find_preventive_hwpx_detail_table_paragraph_index(section_children: Sequence[ET.Element]) -> Optional[int]:
    for child_index, child in enumerate(section_children):
        if child.tag == HP_TAG + "p" and find_preventive_hwpx_detail_table([child]) is not None:
            return child_index
    return None


def remove_hwpx_spacer_paragraphs_after_index(
    section_children: List[ET.Element],
    paragraph_index: Optional[int],
    remove_count: int,
) -> None:
    if paragraph_index is None or remove_count < 1:
        return

    child_index = paragraph_index + 1
    removed_count = 0
    while child_index < len(section_children) and removed_count < remove_count:
        child = section_children[child_index]
        if not is_blank_hwpx_paragraph(child):
            break
        del section_children[child_index]
        removed_count += 1


def is_blank_hwpx_paragraph(element: ET.Element) -> bool:
    if element.tag != HP_TAG + "p":
        return False
    if element.find(".//" + HP_TAG + "tbl") is not None:
        return False
    return all(not (text_element.text or "").strip() for text_element in element.iter(HP_TAG + "t"))


def find_preventive_hwpx_detail_table(section_children: Sequence[ET.Element]) -> Optional[ET.Element]:
    for child in section_children:
        for table in child.iter(HP_TAG + "tbl"):
            if hwpx_element_contains_text(table, "{{NUM}}"):
                return table
    return None


def hwpx_element_contains_text(element: ET.Element, needle: str) -> bool:
    for text_element in element.iter(HP_TAG + "t"):
        if needle in (text_element.text or ""):
            return True
    return False


def replace_hwpx_placeholders(elements: Sequence[ET.Element], replacements: Mapping[str, Any]) -> None:
    for element in elements:
        for text_element in element.iter(HP_TAG + "t"):
            text = text_element.text or ""
            for placeholder, value in replacements.items():
                text = text.replace(placeholder, str(value))
            text_element.text = text


def set_hwpx_table_row_address(row: ET.Element, row_address: int) -> None:
    for cell in row.findall(HP_TAG + "tc"):
        cell_address = cell.find(HP_TAG + "cellAddr")
        if cell_address is not None:
            cell_address.set("rowAddr", str(row_address))


def set_hwpx_table_cell_text(row: ET.Element, column_index: int, value: Any) -> None:
    cells = row.findall(HP_TAG + "tc")
    if column_index >= len(cells):
        return

    cell = cells[column_index]
    text_elements = list(cell.iter(HP_TAG + "t"))
    if text_elements:
        text_elements[0].text = "" if value is None else str(value)
        for extra_text_element in text_elements[1:]:
            extra_text_element.text = ""
        return

    run = cell.find(".//" + HP_TAG + "run")
    if run is not None:
        text_element = ET.SubElement(run, HP_TAG + "t")
        text_element.text = "" if value is None else str(value)


def count_detail_result_statuses(details_for_host: Sequence[DetailRow]) -> Tuple[int, int]:
    pass_count = 0
    fail_count = 0
    for detail_row in details_for_host:
        status_type = classify_detail_result_status(detail_row.result_status)
        if status_type == "pass":
            pass_count += 1
        elif status_type == "fail":
            fail_count += 1
    return pass_count, fail_count


def build_area_document_filename(
    output_stem: str,
    area_name: str,
    section_index: int,
    used_names: Set[str],
    suffix: str = ".docx",
) -> str:
    normalized_suffix = normalize_suffix(suffix)
    cleaned_area = INVALID_FILENAME_CHARS.sub("_", area_name or "").strip().strip(".")
    cleaned_area = re.sub(r"\s+", " ", cleaned_area) or "점검 영역"
    base_name = f"{output_stem}_{section_index:02d}_{cleaned_area}"
    candidate = f"{base_name}{normalized_suffix}"
    dedupe_suffix = 1
    while candidate.lower() in used_names:
        candidate = f"{base_name}_{dedupe_suffix}{normalized_suffix}"
        dedupe_suffix += 1
    used_names.add(candidate.lower())
    return candidate


def build_government_checklist_docx(
    area_name: str,
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
) -> bytes:
    document_xml = build_government_checklist_document_xml(area_name, host_entries)
    package_files = {
        "[Content_Types].xml": build_docx_content_types_xml(),
        "_rels/.rels": build_docx_root_rels_xml(),
        "docProps/core.xml": build_docx_core_properties_xml(),
        "docProps/app.xml": build_docx_app_properties_xml(),
        "word/document.xml": document_xml,
        "word/_rels/document.xml.rels": build_docx_document_rels_xml(),
        "word/styles.xml": build_docx_styles_xml(),
        "word/settings.xml": build_docx_settings_xml(),
        "word/webSettings.xml": build_docx_web_settings_xml(),
        "word/fontTable.xml": build_docx_font_table_xml(),
    }

    from io import BytesIO

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zip_handle:
        for filename, content in package_files.items():
            zip_handle.writestr(filename, content.encode("utf-8"))
    return buffer.getvalue()


def build_government_checklist_document_xml(
    area_name: str,
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
) -> str:
    inspected_area = (area_name or "").strip() or "점검 영역"
    host_labels = [build_host_label(summary_row) for summary_row, _details in host_entries]
    checked_time = format_datetime(first_present_value(*build_government_checklist_checked_time_candidates(host_entries)))
    host_detail_indexes = [
        index_details_by_government_checklist_item(details_for_host) for _summary_row, details_for_host in host_entries
    ]
    item_entries = build_government_checklist_item_entries(host_entries)
    all_details_for_area = [detail for _summary_row, details_for_host in host_entries for detail in details_for_host]

    body_parts = [
        build_docx_paragraph("[서식 1] 정보시스템 상태점검", style="Normal"),
        build_docx_paragraph(f"{inspected_area} 점검 일지", style="Title", align="center"),
        build_docx_paragraph(f"□ {inspected_area} 점검", style="Normal"),
        build_docx_paragraph(f"○ 장비명: {', '.join(host_labels) if host_labels else '-'}", style="Normal"),
        build_docx_paragraph(f"○ 점검 일자: {checked_time}", style="Normal"),
        build_government_checklist_docx_table(item_entries, host_entries, host_detail_indexes),
        build_docx_paragraph(f"□ 점검 결과 요약: {build_government_checklist_result_summary(all_details_for_area)}"),
        build_docx_paragraph(f"□ 주요 조치내역: {build_government_checklist_action_summary(host_entries)}"),
        build_docx_paragraph("점검자: __________ (서명)        확인자: __________ (서명)", align="center"),
    ]
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        + "".join(body_parts)
        + '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="1134" w:right="850" w:bottom="1134" w:left="850" w:header="708" w:footer="708" w:gutter="0"/>'
        "</w:sectPr></w:body></w:document>"
    )


def build_government_checklist_docx_table(
    item_entries: Sequence[Tuple[str, str]],
    host_entries: Sequence[Tuple[SummaryRow, Sequence[DetailRow]]],
    host_detail_indexes: Sequence[Mapping[str, DetailRow]],
) -> str:
    column_count = 2 + (len(host_entries) * 3)
    grid_columns = ['<w:gridCol w:w="650"/>', '<w:gridCol w:w="2600"/>']
    for _summary_row, _details_for_host in host_entries:
        grid_columns.extend(
            ['<w:gridCol w:w="720"/>', '<w:gridCol w:w="820"/>', '<w:gridCol w:w="1550"/>']
        )

    rows = [
        "<w:tr>"
        + build_docx_table_cell("번호", v_merge="restart", width=650, align="center", bold=True)
        + build_docx_table_cell("점검항목", v_merge="restart", width=2600, align="center", bold=True)
        + "".join(
            build_docx_table_cell(build_host_label(summary_row), grid_span=3, align="center", bold=True)
            for summary_row, _details_for_host in host_entries
        )
        + "</w:tr>",
        "<w:tr>"
        + build_docx_table_cell("", v_merge="continue", width=650)
        + build_docx_table_cell("", v_merge="continue", width=2600)
        + "".join(
            build_docx_table_cell(header, align="center", bold=True)
            for _summary_row, _details_for_host in host_entries
            for header in ("정상", "비정상", "비고")
        )
        + "</w:tr>",
    ]

    if not item_entries:
        rows.append(
            "<w:tr>"
            + build_docx_table_cell("상세 데이터가 없습니다.", grid_span=max(column_count, 1))
            + "</w:tr>"
        )
    else:
        for item_number, (item_key, item_label) in enumerate(item_entries, start=1):
            result_cells = []
            for host_detail_index in host_detail_indexes:
                detail_row = host_detail_index.get(item_key)
                result_cells.extend(build_government_checklist_result_cells(detail_row))
            rows.append(
                "<w:tr>"
                + build_docx_table_cell(str(item_number), align="center")
                + build_docx_table_cell(item_label)
                + "".join(
                    build_docx_table_cell(
                        value,
                        align="center" if index % 3 in {0, 1} else "left",
                    )
                    for index, value in enumerate(result_cells)
                )
                + "</w:tr>"
            )

    return (
        '<w:tbl><w:tblPr><w:tblW w:w="0" w:type="auto"/>'
        '<w:tblBorders><w:top w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:left w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:right w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:insideH w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:insideV w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        "</w:tblBorders></w:tblPr>"
        f"<w:tblGrid>{''.join(grid_columns)}</w:tblGrid>"
        + "".join(rows)
        + "</w:tbl>"
    )


def build_docx_table_cell(
    text: Any,
    *,
    grid_span: int = 1,
    v_merge: Optional[str] = None,
    width: Optional[int] = None,
    align: str = "left",
    bold: bool = False,
) -> str:
    properties = []
    if width:
        properties.append(f'<w:tcW w:w="{int(width)}" w:type="dxa"/>')
    if grid_span > 1:
        properties.append(f'<w:gridSpan w:val="{int(grid_span)}"/>')
    if v_merge == "restart":
        properties.append('<w:vMerge w:val="restart"/>')
    elif v_merge == "continue":
        properties.append("<w:vMerge/>")
    return f"<w:tc><w:tcPr>{''.join(properties)}</w:tcPr>{build_docx_paragraph(text, align=align, bold=bold)}</w:tc>"


def build_docx_paragraph(
    text: Any,
    *,
    style: str = "Normal",
    align: str = "left",
    bold: bool = False,
) -> str:
    style_xml = f'<w:pStyle w:val="{xml_escape(style)}"/>' if style != "Normal" else ""
    align_xml = f'<w:jc w:val="{xml_escape(align)}"/>' if align else ""
    bold_xml = "<w:b/>" if bold else ""
    text_xml = build_docx_text_runs(text)
    return (
        f"<w:p><w:pPr>{style_xml}{align_xml}</w:pPr>"
        f'<w:r><w:rPr><w:rFonts w:ascii="Malgun Gothic" w:eastAsia="Malgun Gothic" w:hAnsi="Malgun Gothic"/>{bold_xml}</w:rPr>'
        f"{text_xml}</w:r></w:p>"
    )


def build_docx_text_runs(text: Any) -> str:
    lines = str(text if text is not None else "").splitlines() or [""]
    parts = []
    for line_index, line in enumerate(lines):
        if line_index:
            parts.append("<w:br/>")
        parts.append(f'<w:t xml:space="preserve">{xml_escape(line)}</w:t>')
    return "".join(parts)


def xml_escape(value: Any) -> str:
    return html.escape(str(value if value is not None else ""), quote=True)


def build_docx_content_types_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
        '<Override PartName="/word/settings.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.settings+xml"/>'
        '<Override PartName="/word/webSettings.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.webSettings+xml"/>'
        '<Override PartName="/word/fontTable.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.fontTable+xml"/>'
        "</Types>"
    )


def build_docx_root_rels_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        "</Relationships>"
    )


def build_docx_document_rels_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/settings" Target="settings.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/webSettings" Target="webSettings.xml"/>'
        '<Relationship Id="rId4" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/fontTable" Target="fontTable.xml"/>'
        "</Relationships>"
    )


def build_docx_core_properties_xml() -> str:
    generated_at = datetime.now().isoformat(timespec="seconds") + "Z"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        "<dc:title>점검보고서</dc:title>"
        "<dc:creator>FAP Report Generator</dc:creator>"
        "<cp:lastModifiedBy>FAP Report Generator</cp:lastModifiedBy>"
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{generated_at}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{generated_at}</dcterms:modified>'
        "</cp:coreProperties>"
    )


def build_docx_app_properties_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        "<Application>FAP Report Generator</Application>"
        "<DocSecurity>0</DocSecurity>"
        "<ScaleCrop>false</ScaleCrop>"
        "<Company></Company>"
        "<LinksUpToDate>false</LinksUpToDate>"
        "<SharedDoc>false</SharedDoc>"
        "<HyperlinksChanged>false</HyperlinksChanged>"
        "<AppVersion>1.0</AppVersion>"
        "</Properties>"
    )


def build_docx_settings_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:settings xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:zoom w:percent="100"/>'
        '<w:defaultTabStop w:val="720"/>'
        '<w:compat><w:compatSetting w:name="compatibilityMode" '
        'w:uri="http://schemas.microsoft.com/office/word" w:val="15"/></w:compat>'
        "</w:settings>"
    )


def build_docx_web_settings_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:webSettings xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:optimizeForBrowser/>"
        "</w:webSettings>"
    )


def build_docx_font_table_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:fonts xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:font w:name="Malgun Gothic"><w:family w:val="swiss"/><w:pitch w:val="variable"/></w:font>'
        "</w:fonts>"
    )


def build_docx_styles_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:style w:type="paragraph" w:default="1" w:styleId="Normal">'
        '<w:name w:val="Normal"/><w:rPr><w:rFonts w:ascii="Malgun Gothic" w:eastAsia="Malgun Gothic" w:hAnsi="Malgun Gothic"/><w:sz w:val="20"/></w:rPr>'
        "</w:style>"
        '<w:style w:type="paragraph" w:styleId="Title">'
        '<w:name w:val="Title"/><w:basedOn w:val="Normal"/><w:rPr><w:b/><w:sz w:val="32"/></w:rPr>'
        "</w:style>"
        "</w:styles>"
    )


def build_result_payload(
    *,
    result: str,
    report_path: Optional[str],
    msg: str,
    job_id: int,
    report_type: str,
    user_id: Optional[str] = None,
):
    payload = {
        "result": result,
        "report_path": report_path,
        "msg": msg,
        "job_id": job_id,
        "report_type": report_type,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    if user_id:
        payload["user_id"] = user_id
    return payload


def format_metric(value: float) -> str:
    rounded = round(float(value), 2)
    return str(int(rounded)) if rounded.is_integer() else f"{rounded:.2f}"


def format_datetime(value: Any) -> str:
    if value in (None, ""):
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def format_duration(seconds: Optional[int]) -> str:
    if seconds in (None, ""):
        return "-"
    total_seconds = max(int(seconds), 0)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def autosize_sheet(
    worksheet: Any,
    get_column_letter: Any,
    fixed_widths: Optional[Mapping[str, float]] = None,
) -> None:
    fixed_width_map = dict(fixed_widths or {})
    for column_cells in worksheet.columns:
        first_cell = column_cells[0]
        column_letter = get_column_letter(first_cell.column)
        if column_letter in fixed_width_map:
            worksheet.column_dimensions[column_letter].width = fixed_width_map[column_letter]
            continue

        max_length = 0
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 30)


def write_label_value_pairs_row(
    worksheet: Any,
    *,
    row: int,
    pairs: Sequence[Tuple[int, int, int, str, Any]],
    label_font: Any,
    value_font: Any,
    label_fill: Any,
    value_fill: Any,
    border: Any,
    label_alignment: Any,
    value_alignment: Any,
) -> None:
    for label_column, value_start_column, value_end_column, label, value in pairs:
        label_cell = worksheet.cell(row=row, column=label_column, value=label)
        label_cell.font = label_font
        label_cell.fill = label_fill
        label_cell.border = border
        label_cell.alignment = label_alignment

        for column_index in range(value_start_column, value_end_column + 1):
            cell = worksheet.cell(row=row, column=column_index)
            cell.fill = value_fill
            cell.border = border
            cell.alignment = value_alignment

        if value_start_column < value_end_column:
            worksheet.merge_cells(
                start_row=row,
                start_column=value_start_column,
                end_row=row,
                end_column=value_end_column,
            )

        value_cell = worksheet.cell(row=row, column=value_start_column, value=value)
        value_cell.font = value_font
        value_cell.fill = value_fill
        value_cell.border = border
        value_cell.alignment = value_alignment


def estimate_text_row_height(
    value: Any,
    *,
    chars_per_line: int = 80,
    line_height: float = 18.0,
    min_height: float = 36.0,
    max_height: float = 180.0,
) -> float:
    text = "" if value in (None, "") else str(value)
    if not text:
        return min_height

    visual_line_count = 0
    for raw_line in text.splitlines() or [""]:
        line_length = len(raw_line)
        visual_line_count += max(1, (line_length + chars_per_line - 1) // chars_per_line)

    return min(max(visual_line_count * line_height, min_height), max_height)


def average_or_zero(values: Iterable[Any]) -> float:
    collected = [float(value) for value in values]
    return mean(collected) if collected else 0.0


def format_importance(value: Any) -> str:
    if value in (None, ""):
        return "-"

    if isinstance(value, (int, float)):
        numeric_value = int(value)
    else:
        text = str(value).strip()
        try:
            numeric_value = int(float(text))
        except ValueError:
            return text or "-"

    return {1: "하", 2: "중", 3: "상"}.get(numeric_value, str(numeric_value))


def normalize_output_name(output_name: str) -> str:
    candidate = Path(output_name).name.strip()
    if not candidate:
        return "점검보고서"

    suffix = Path(candidate).suffix
    if suffix.lower() == ".xlsx":
        candidate = Path(candidate).stem

    candidate = candidate.strip()
    return candidate or "점검보고서"


def _to_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    return int(value)


def _to_optional_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    return int(value)


def _to_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    return float(value)


def _to_optional_bool(value: Any) -> Optional[bool]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.lower()
        if lowered in {"true", "t", "1", "y", "yes"}:
            return True
        if lowered in {"false", "f", "0", "n", "no"}:
            return False
    return bool(value)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    resolved_user_id = resolve_user_id(args.user_id)

    try:
        if args.mock_host_count > 0:
            summary_rows, detail_rows = build_mock_report_rows(
                args.job_id,
                args.mock_host_count,
                args.mock_items_per_host,
            )
        else:
            api_url, api_token = load_api_config()
            summary_rows, detail_rows = fetch_report_rows(api_url, api_token, args.job_id, resolved_user_id)
        if not summary_rows:
            payload = build_result_payload(
                result="error",
                report_path=None,
                msg="no data",
                job_id=args.job_id,
                report_type=args.report_type,
                user_id=resolved_user_id,
            )
            print(json.dumps(payload, ensure_ascii=False))
            return 1

        if args.report_type == "preventive":
            output_suffix = ".hwpx"
        elif args.output_format == "docx":
            output_suffix = ".docx"
        else:
            output_suffix = ".xlsx"
        if args.file_name:
            output_path = build_named_output_path(args.file_name, args.output_path, args.output_dir, output_suffix)
        elif args.output_path:
            output_path = build_explicit_output_path(args.output_path, output_suffix)
        else:
            output_path = build_output_path(args.output_name, args.output_dir, output_suffix)
        msg = "success"
        if args.report_type == "preventive":
            primary_report_path = str(save_preventive_hwpx_reports(summary_rows, detail_rows, output_path))
        elif args.output_format == "docx":
            primary_report_path = str(save_government_checklist_docx_reports(summary_rows, detail_rows, output_path))
        else:
            generator = get_report_generator(args.report_type)
            saved_paths = generator.save_workbooks(summary_rows, detail_rows, output_path)
            primary_report_path = str(saved_paths[0])
            if len(saved_paths) > 1:
                archive_path = build_zip_archive(saved_paths, build_zip_output_path(output_path))
                primary_report_path = str(archive_path)

        payload = build_result_payload(
            result="success",
            report_path=primary_report_path,
            msg=msg,
            job_id=args.job_id,
            report_type=args.report_type,
            user_id=resolved_user_id,
        )
        print(json.dumps(payload, ensure_ascii=False))
        return 0
    except Exception as exc:
        error_text = str(exc).lower()
        if "status=401" in error_text or "unauthorized" in error_text:
            msg = "unauthorized"
        elif "api request failed" in error_text or "invalid json response from api" in error_text:
            msg = "api error"
        else:
            msg = "report create failed"
        payload = build_result_payload(
            result="error",
            report_path=None,
            msg=msg,
            job_id=args.job_id,
            report_type=args.report_type,
            user_id=resolved_user_id,
        )
        print(json.dumps(payload, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    sys.exit(main())
